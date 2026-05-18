"""
Accounts app views

Key areas:
- Authentication and profile management (register, login, profile, stats)
- Organization structure (branches, departments, units, corpers)
- Attendance capture with geofence and simple face detection
- Performance clearance generation and verification
- Wallets and transactions (ORG/BRANCH/CORPER) with deduction order:
  ORG -> BRANCH -> CORPER; welcome bonus applies to ORG only
- System settings, announcements for org dashboards
- Paystack funding: initialize and verify using keys stored in DB

Notes:
- Uses timezone-aware `timezone.localtime()` for logging and `Africa/Lagos` by default
- Avoids double charges by checking reference across wallets
"""

from django.contrib.auth import get_user_model, authenticate, login, logout
from django.conf import settings
from django.shortcuts import redirect
from rest_framework import status, permissions, viewsets
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from django.middleware.csrf import get_token
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from django.core.cache import cache
from django.http import JsonResponse, HttpResponseNotAllowed, HttpResponseNotFound
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
import requests
import base64
import csv
import math
import numpy as np
import cv2
import os
import re
import shutil
import uuid
import traceback
import face_recognition
import hmac
import hashlib
import json
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO

from .serializers import (
    OrganizationRegisterSerializer,
    LoginSerializer,
    OrganizationProfileSerializer,
    BranchOfficeSerializer,
    DepartmentSerializer,
    UnitSerializer,
    CorpMemberSerializer,
    PublicHolidaySerializer,
    LeaveRequestSerializer,
    NotificationSerializer,
    QueryRecordSerializer,
)
from .tokens import validate_email_token, generate_email_token
from .models import OrganizationProfile, BranchOffice, Department, Unit, CorpMember, PublicHoliday, LeaveRequest, Notification, AttendanceLog, WalletAccount, WalletTransaction, ClearanceOverride, TempFaceEncoding
from .models import QueryRecord
from .models import NationalHoliday
from .models import SubscriptionPlanSetting, OrganizationSubscription, SubscriptionPayment, ClearanceAccess, SystemSetting, PaystackConfig
from .services.holidays import ensure_national_holidays, is_holiday_for_org, working_days
from django.db.models import Count
from django.db import models, transaction


User = get_user_model()

# In-memory capture state per corper_id
_CAPTURE_STATE = {}
_ATTENDANCE_STATE = {}

def _ensure_rgb_uint8(arr):
    """Ensure image array is uint8 RGB (H, W, 3).
    - Cast dtype to uint8 (with clipping).
    - Drop alpha channel if present.
    - Expand/convert grayscale to 3 channels.
    """
    if arr is None:
        return None
    try:
        # Ensure dtype uint8; scale floats in 0..1 range if needed
        if arr.dtype != np.uint8:
            if np.issubdtype(arr.dtype, np.floating):
                scale = 255.0 if float(np.nanmax(arr) or 0.0) <= 1.0 else 1.0
                arr = np.clip(arr * scale, 0, 255).astype(np.uint8)
            else:
                arr = np.clip(arr, 0, 255).astype(np.uint8)
        # Ensure 3 channels
        if arr.ndim == 2:
            arr = cv2.cvtColor(arr, cv2.COLOR_GRAY2RGB)
        elif arr.ndim == 3:
            if arr.shape[2] == 4:
                arr = arr[:, :, :3]
            elif arr.shape[2] == 3:
                pass
            else:
                # Unexpected channel count; best effort: take first 3 or tile
                if arr.shape[2] > 3:
                    arr = arr[:, :, :3]
                else:
                    arr = np.repeat(arr, 3, axis=2)[:, :, :3]
        return arr
    except Exception as e:
        print("[capture] _ensure_rgb_uint8 error:", e)
        traceback.print_exc()
        return arr

def sanitize_rgb(img):
    """Convert a decoded frame to uint8 3-channel RGB.
    - Accepts BGR/GRAY/BGRA inputs, drops alpha if present.
    - Casts dtype to uint8 with clipping/scaling when necessary.
    """
    if img is None:
        return None
    try:
        arr = img
        # Type normalize first
        if arr.dtype != np.uint8:
            if np.issubdtype(arr.dtype, np.floating):
                maxv = float(np.nanmax(arr)) if arr.size else 1.0
                scale = 255.0 if maxv <= 1.0 else 1.0
                arr = np.clip(arr * scale, 0, 255).astype(np.uint8)
            else:
                arr = np.clip(arr, 0, 255).astype(np.uint8)
        # Channel handling and BGR->RGB conversion
        if arr.ndim == 2:
            rgb = cv2.cvtColor(arr, cv2.COLOR_GRAY2RGB)
        elif arr.ndim == 3:
            ch = arr.shape[2]
            if ch == 4:
                arr = arr[:, :, :3]
            if arr.shape[2] >= 3:
                rgb = cv2.cvtColor(arr[:, :, :3], cv2.COLOR_BGR2RGB)
            else:
                rgb = cv2.cvtColor(arr, cv2.COLOR_GRAY2RGB)
        else:
            rgb = None
        rgb = _ensure_rgb_uint8(rgb) if rgb is not None else None
        if rgb is None:
            return None
        if not rgb.flags['C_CONTIGUOUS']:
            rgb = np.ascontiguousarray(rgb)
        if rgb.dtype != np.uint8:
            rgb = rgb.astype(np.uint8, copy=False)
        if rgb.ndim != 3 or rgb.shape[2] != 3:
            return None
        return rgb
    except Exception as e:
        print("[capture] sanitize_rgb error:", e)
        traceback.print_exc()
        return None

def ensure_dlib_rgb(arr):
    """Return a new C-contiguous uint8 RGB array suitable for dlib.
    Always makes a copy to avoid any exotic strides/views that dlib rejects.
    """
    if arr is None:
        return None
    try:
        a = arr
        if a.ndim == 2:
            a = cv2.cvtColor(a, cv2.COLOR_GRAY2RGB)
        elif a.ndim == 3:
            if a.shape[2] == 4:
                a = a[:, :, :3]
            elif a.shape[2] < 3:
                a = cv2.cvtColor(a, cv2.COLOR_GRAY2RGB)
        else:
            return None
        # Enforce dtype and memory layout with a copy
        a = np.array(a, dtype=np.uint8, order='C', copy=True)
        return a
    except Exception as e:
        print("[capture] ensure_dlib_rgb error:", e)
        traceback.print_exc()
        return None

def _safe_face_locations(img_rgb_or_gray, tag=""):
    """Call face_recognition.face_locations safely on a valid uint8 image.
    - First try on the provided image (expected RGB from sanitize_rgb).
    - If dlib complains about image type, fall back to grayscale detection.
    """
    try:
        locs = face_recognition.face_locations(img_rgb_or_gray)
        if not locs:
            # Try a higher upsample for small frames
            locs = face_recognition.face_locations(img_rgb_or_gray, number_of_times_to_upsample=2)
        if locs:
            return locs
        # If empty (no exception), try grayscale HOG as a second strategy
        arr = img_rgb_or_gray
        try:
            if arr is None:
                return []
            if arr.ndim == 3 and arr.shape[2] == 3:
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
            elif arr.ndim == 2:
                gray = arr
            else:
                return []
            if gray.dtype != np.uint8:
                gray = np.clip(gray, 0, 255).astype(np.uint8)
            if not gray.flags['C_CONTIGUOUS']:
                gray = np.ascontiguousarray(gray)
            _debug_img(f'{tag}.gray_noerr', gray)
            locs = face_recognition.face_locations(gray)
            if not locs:
                locs = face_recognition.face_locations(gray, number_of_times_to_upsample=2)
            if locs:
                return locs
        except Exception as ee_g:
            print(f"[detect] grayscale no-error path failed on {tag}: {ee_g}")
            traceback.print_exc()
        # Final fallback: Haar cascade
        try:
            face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
            locs = []
            for (x, y, w, h) in faces:
                top = int(y)
                left = int(x)
                bottom = int(y + h)
                right = int(x + w)
                locs.append((top, right, bottom, left))
            print(f"[detect] Haar fallback (no-error) produced {len(locs)} faces for {tag}")
            return locs
        except Exception as ee3:
            print(f"[detect] Haar fallback (no-error) failed on {tag}: {ee3}")
            traceback.print_exc()
            return []
    except RuntimeError as e:
        msg = str(e)
        print(f"[detect] face_locations failed on {tag}: {msg}")
        traceback.print_exc()
        # Fallback to grayscale if possible
        try:
            arr = img_rgb_or_gray
            if arr is None:
                return []
            if arr.ndim == 3 and arr.shape[2] == 3:
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
            elif arr.ndim == 2:
                gray = arr
            else:
                return []
            if gray.dtype != np.uint8:
                gray = np.clip(gray, 0, 255).astype(np.uint8)
            if not gray.flags['C_CONTIGUOUS']:
                gray = np.ascontiguousarray(gray)
            _debug_img(f'{tag}.gray', gray)
            try:
                locs = face_recognition.face_locations(gray)
                if not locs:
                    locs = face_recognition.face_locations(gray, number_of_times_to_upsample=2)
                if not locs:
                    print(f"[detect] No faces with HOG on gray (upsample<=2) for {tag}")
                return locs
            except RuntimeError as ee2:
                print(f"[detect] face_locations failed on grayscale {tag}: {ee2}")
                traceback.print_exc()
                # Final fallback: Haar cascade via OpenCV
                try:
                    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
                    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
                    locs = []
                    for (x, y, w, h) in faces:
                        top = int(y)
                        left = int(x)
                        bottom = int(y + h)
                        right = int(x + w)
                        locs.append((top, right, bottom, left))
                    print(f"[detect] Haar fallback produced {len(locs)} faces for {tag}")
                    return locs
                except Exception as ee3:
                    print(f"[detect] Haar fallback failed on {tag}: {ee3}")
                    traceback.print_exc()
                    return []
        except Exception as ee:
            print(f"[detect] grayscale fallback failed on {tag}: {ee}")
            traceback.print_exc()
            return []
    except Exception as e:
        print(f"[detect] face_locations unexpected error on {tag}: {e}")
        traceback.print_exc()
        return []
    

def _reset_capture_state(corper_id: int):
    _CAPTURE_STATE[corper_id] = {
        'enc_count': 0,   # number of encodings accumulated
        'enc_sum': None,  # running sum of encoding vectors (numpy array)
    }

def _reset_attendance_state(corper_id: int):
    _ATTENDANCE_STATE[corper_id] = {
        'hits': 0,
        'logged': False,
    }

def _debug_img(tag, arr):
    try:
        if arr is None:
            print(f"[debug] {tag}: arr=None")
            return
        print(f"[debug] {tag}: dtype={arr.dtype}, shape={getattr(arr, 'shape', None)}, contig={arr.flags['C_CONTIGUOUS']}, strides={arr.strides}")
        try:
            amin = float(np.min(arr))
            amax = float(np.max(arr))
            print(f"[debug] {tag}: min={amin} max={amax}")
        except Exception:
            pass
    except Exception as e:
        print(f"[debug] {tag}: error printing info: {e}")
        traceback.print_exc()

def _process_capture_frame(corper_id: int, b64_frame: str, save_dir: str):
    # save_dir no longer used; kept for signature compatibility
    st = _CAPTURE_STATE.setdefault(corper_id, {'enc_count': 0, 'enc_sum': None})
    img_data = base64.b64decode(b64_frame)
    nparr = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        return None, st['enc_count']

    # Sanitize to contiguous uint8 RGB for face_recognition
    rgb = sanitize_rgb(img)
    # Detect faces and compute first encoding
    try:
        locations = _safe_face_locations(rgb, tag='capture_inmem')
        img_for_enc = ensure_dlib_rgb(rgb)
        _debug_img('capture_inmem.enc_img', img_for_enc)
        encs = face_recognition.face_encodings(img_for_enc, locations) if img_for_enc is not None else []
    except Exception as e:
        print("[capture] Error in face_encodings (_process_capture_frame):", e)
        traceback.print_exc()
        locations, encs = [], []

    # If we have an encoding and haven't reached 30 yet, update running sum
    MAX_ENCODINGS = 30
    if encs and st.get('enc_count', 0) < MAX_ENCODINGS:
        vec = np.asarray(encs[0], dtype=np.float32)
        if st.get('enc_sum') is None:
            st['enc_sum'] = vec.copy()
        else:
            st['enc_sum'] = st['enc_sum'] + vec
        st['enc_count'] = st.get('enc_count', 0) + 1

    # draw overlays (use first face location if present)
    if locations:
        top, right, bottom, left = locations[0]
        cv2.rectangle(img, (left, top), (right, bottom), (0,255,0), 2)
        cv2.putText(
            img,
            f'Enc {st.get("enc_count",0)}/{MAX_ENCODINGS}',
            (left, max(0, top-10)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (0,255,0),
            2,
            cv2.LINE_AA
        )
    _, buffer = cv2.imencode('.jpg', img)
    processed_frame = base64.b64encode(buffer).decode('utf-8')
    return processed_frame, st.get('enc_count', 0)


def capture_page(request, corper_id: int):
    # Authorization: ORG or BRANCH can access; corper can also view own page
    user = request.user
    try:
        corper = CorpMember.objects.get(pk=corper_id)
    except CorpMember.DoesNotExist:
        return HttpResponseNotFound('Corper not found')

    # Basic auth checks
    allowed = False
    if getattr(user, 'role', None) == 'ORG' and getattr(corper.user, 'id', None) == user.id:
        allowed = True
    if getattr(user, 'role', None) == 'BRANCH':
        b = BranchOffice.objects.filter(admin=user).first()
        if b and b.id == getattr(corper.branch, 'id', None):
            allowed = True
    if getattr(user, 'role', None) == 'CORPER' and getattr(corper.account, 'id', None) == user.id:
        allowed = True
    if not allowed:
        raise PermissionDenied('Not allowed')

    _reset_capture_state(corper_id)
    # Prefer request Origin if it matches configured FRONTEND_ORIGINS
    request_origin = request.headers.get('Origin')
    try:
        allowed = set(getattr(settings, 'FRONTEND_ORIGINS', []))
    except Exception:
        allowed = set()
    frontend_base = (request_origin if request_origin in allowed else getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173'))).rstrip('/')

    # Ensure CSRF cookie is set for subsequent POSTs from the page
    try:
        get_token(request)
    except Exception:
        pass

    # Generate a unique session id for this capture flow
    session_id = uuid.uuid4().hex

    ctx = {
        'corper_id': corper.id,
        'full_name': corper.full_name,
        'state_code': corper.state_code,
        'frontend_base': frontend_base,
        'session_id': session_id,
    }
    return render(request, 'capture.html', ctx)


def capture_process_frame(request, corper_id: int):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    frame = request.POST.get('frame')
    session_id = request.POST.get('session')
    if not frame:
        return JsonResponse({'detail': 'Missing frame'}, status=400)
    if not session_id:
        return JsonResponse({'detail': 'Missing session'}, status=400)
    # Save dir per corper under media/captures/{id}
    media_root = getattr(settings, 'MEDIA_ROOT', os.path.join(os.getcwd(), 'media'))
    save_dir = os.path.join(media_root, 'captures', str(corper_id))
    processed, _ = _process_capture_frame(corper_id, frame, save_dir)
    if not processed:
        return JsonResponse({'detail': 'Processing failed'}, status=500)

    # Compute encoding for persistence (stateless across workers)
    try:
        img_data = base64.b64decode(frame)
        nparr = np.frombuffer(img_data, np.uint8)
        bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        rgb = sanitize_rgb(bgr)
        _debug_img('capture_process_frame.rgb', rgb)
        locs = _safe_face_locations(rgb, tag='capture_persist') if rgb is not None else []
        img_for_enc = ensure_dlib_rgb(rgb) if rgb is not None else None
        _debug_img('capture_persist.enc_img', img_for_enc)
        encs = face_recognition.face_encodings(img_for_enc, locs) if img_for_enc is not None else []
    except Exception as e:
        print("[capture] Error in face_locations/encodings (persist step):", e)
        traceback.print_exc()
        encs = []

    MAX_ENCODINGS = 30
    current_count = TempFaceEncoding.objects.filter(corper_id=corper_id, session_id=session_id).count()
    if encs and current_count < MAX_ENCODINGS:
        import json
        vec = np.asarray(encs[0], dtype=np.float32)
        TempFaceEncoding.objects.create(
            corper_id=corper_id,
            session_id=session_id,
            idx=current_count,
            vector=json.dumps(vec.tolist()),
        )
        current_count += 1

    return JsonResponse({'frame': processed, 'saved': current_count})


def capture_finalize(request, corper_id: int):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    # Auth similar to capture_page
    user = request.user
    try:
        corper = CorpMember.objects.get(pk=corper_id)
    except CorpMember.DoesNotExist:
        return HttpResponseNotFound('Corper not found')

    allowed = False
    if getattr(user, 'role', None) == 'ORG' and getattr(corper.user, 'id', None) == user.id:
        allowed = True
    if getattr(user, 'role', None) == 'BRANCH':
        b = BranchOffice.objects.filter(admin=user).first()
        if b and b.id == getattr(corper.branch, 'id', None):
            allowed = True
    if getattr(user, 'role', None) == 'CORPER' and getattr(corper.account, 'id', None) == user.id:
        allowed = True
    if not allowed:
        return JsonResponse({'detail': 'Not allowed'}, status=403)

    session_id = request.POST.get('session')
    if not session_id:
        return JsonResponse({'detail': 'Missing session'}, status=400)

    # Collect encodings from DB for this session
    import json
    rows = list(TempFaceEncoding.objects.filter(corper=corper, session_id=session_id).order_by('idx', 'id'))
    if not rows:
        return JsonResponse({'detail': 'No encodings found; ensure face is visible'}, status=400)
    try:
        arrs = [np.array(json.loads(r.vector), dtype=np.float32) for r in rows]
        avg = np.mean(np.stack(arrs, axis=0), axis=0)
    except Exception:
        return JsonResponse({'detail': 'Failed to build average encoding'}, status=500)
    # Save to model as JSON list of floats
    try:
        corper.face_encoding = json.dumps(avg.tolist())
        corper.save(update_fields=['face_encoding'])
    except Exception as e:
        return JsonResponse({'detail': f'Failed to save encoding: {e}'}, status=500)
    # Delete temp rows for this session
    TempFaceEncoding.objects.filter(corper=corper, session_id=session_id).delete()

    # Reset in-memory state
    _reset_capture_state(corper_id)

    return JsonResponse({'status': 'ok', 'encodings': len(rows)})


def attendance_page(request):
    # Only corpers can mark attendance for themselves
    user = request.user
    if getattr(user, 'role', None) != 'CORPER':
        raise PermissionDenied('Only corpers can access attendance')
    cm = getattr(user, 'corper_profile', None)
    if not cm:
        return HttpResponseNotFound('Corper profile not found')

    _reset_attendance_state(cm.id)

    # Prefer request Origin if it matches configured FRONTEND_ORIGINS
    request_origin = request.headers.get('Origin')
    try:
        allowed = set(getattr(settings, 'FRONTEND_ORIGINS', []))
    except Exception:
        allowed = set()
    frontend_base = (request_origin if request_origin in allowed else getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173'))).rstrip('/')

    # Ensure CSRF cookie is set for JS POSTs from this page
    try:
        get_token(request)
    except Exception:
        pass

    ctx = {
        'corper_id': cm.id,
        'full_name': cm.full_name,
        'state_code': cm.state_code,
        'frontend_base': frontend_base,
    }
    return render(request, 'attendance.html', ctx)


def _get_target_location_for_corper(cm):
    """Return (lat, lng, source) using the strictest available center.
    Strict policy: use branch coordinates if set; otherwise fallback to organization coordinates.
    """
    try:
        br = cm.branch
        if br and br.latitude is not None and br.longitude is not None:
            return br.latitude, br.longitude, 'branch'
    except Exception:
        pass
    try:
        prof = OrganizationProfile.objects.filter(user=cm.user).first()
        if prof and prof.location_lat is not None and prof.location_lng is not None:
            return prof.location_lat, prof.location_lng, 'org'
    except Exception:
        pass
    return None, None, None


def _haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi/2.0)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dl/2.0)**2
    c = 2*math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R*c


def attendance_authorize(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    user = request.user
    if getattr(user, 'role', None) != 'CORPER':
        return JsonResponse({'detail': 'Not allowed'}, status=403)
    cm = getattr(user, 'corper_profile', None)
    if not cm:
        return JsonResponse({'detail': 'Corper profile not found'}, status=404)
    try:
        lat = float(request.POST.get('lat'))
        lng = float(request.POST.get('lng'))
    except Exception:
        return JsonResponse({'detail': 'Invalid coordinates'}, status=400)
    tgt_lat, tgt_lng, src = _get_target_location_for_corper(cm)
    if tgt_lat is None or tgt_lng is None:
        return JsonResponse({'allowed': False, 'detail': 'Organization location not configured'}, status=400)
    # Threshold (meters); can be tuned or moved to settings
    threshold = getattr(settings, 'ATTENDANCE_GEOFENCE_METERS', 250)
    dist = _haversine_m(lat, lng, tgt_lat, tgt_lng)
    allowed = dist <= threshold
    # Block on public holidays (manual org + national)
    today = timezone.localdate()
    if is_holiday_for_org(cm.user, today).is_holiday:
        return JsonResponse({'allowed': False, 'detail': 'Today is a public holiday for your organization'}, status=403)
    return JsonResponse({'allowed': allowed, 'distance_m': round(dist, 1), 'threshold_m': threshold, 'source': src or 'unknown'})


def attendance_process_frame(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    user = request.user
    if getattr(user, 'role', None) != 'CORPER':
        return JsonResponse({'ok': False, 'reason': 'not_allowed'})
    cm = getattr(user, 'corper_profile', None)
    if not cm:
        return JsonResponse({'ok': False, 'reason': 'no_profile'})
    frame = request.POST.get('frame')
    if not frame:
        return JsonResponse({'ok': False, 'reason': 'missing_frame'})

    # Load stored encoding
    import json
    try:
        saved = np.array(json.loads(cm.face_encoding), dtype='float32') if cm.face_encoding else None
    except Exception:
        saved = None
    if saved is None:
        return JsonResponse({'ok': False, 'reason': 'no_encoding'})

    # Decode incoming frame and detect face(s)
    img_data = base64.b64decode(frame)
    nparr = np.frombuffer(img_data, np.uint8)
    bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if bgr is None:
        return JsonResponse({'ok': False, 'reason': 'invalid_frame'})
    rgb = sanitize_rgb(bgr)
    _debug_img('attendance_process_frame.rgb', rgb)
    # Skip blank/dark frames early
    try:
        if rgb is None or float(np.max(rgb)) <= 5.0:
            return JsonResponse({'ok': False, 'reason': 'blank'})
    except Exception:
        return JsonResponse({'ok': False, 'reason': 'invalid_frame'})
    # Detect faces and compute encodings with locations to draw overlays
    try:
        face_locations = _safe_face_locations(rgb, tag='attendance')
        print(f"[attendance] detected faces: {len(face_locations)}")
        img_for_enc = ensure_dlib_rgb(rgb)
        _debug_img('attendance.enc_img', img_for_enc)
        encs = face_recognition.face_encodings(img_for_enc, face_locations) if img_for_enc is not None else []
    except Exception as e:
        print("[attendance] Error in face_encodings:", e)
        traceback.print_exc()
        face_locations, encs = [], []

    recognized = False
    best_conf = 0.0
    best_idx = -1
    # Iterate faces, compute distance and draw rectangles + confidence bar
    for idx, ((top, right, bottom, left), enc) in enumerate(zip(face_locations, encs)):
        dist = float(face_recognition.face_distance([saved], enc)[0])
        # Convert to 0-100 confidence; clamp
        conf = max(0.0, min(100.0, (1.0 - dist) * 100.0))
        is_match = dist <= 0.6
        # Track best
        if conf > best_conf:
            best_conf = conf
            best_idx = idx
        # Rectangle around face
        cv2.rectangle(bgr, (left, top), (right, bottom), (0, 255, 0) if is_match else (0, 0, 255), 2)
        # Grey background bar
        cv2.rectangle(bgr, (right + 10, top), (right + 30, bottom), (128, 128, 128), -1)
        # Filled confidence portion (green if >=65 else red)
        bar_height = bottom - top
        filled = int(bar_height * (conf / 100.0))
        color = (0, 255, 0) if conf >= 65 else (0, 0, 255)
        cv2.rectangle(bgr, (right + 10, bottom - filled), (right + 30, bottom), color, -1)
        # If confident, show name + state_code box
        if conf >= 65:
            cv2.rectangle(bgr, (right + 35, top), (right + 260, top + 45), (128, 128, 128), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(bgr, f"{cm.full_name}", (right + 40, top + 40), font, 0.6, (0, 255, 0), 1)
            cv2.putText(bgr, f"{cm.state_code}", (right + 40, top + 20), font, 0.6, (0, 255, 0), 1)

    # Update recognition state using the best face confidence
    st = _ATTENDANCE_STATE.setdefault(cm.id, {'hits': 0, 'logged': False})
    if best_conf >= 65:  # ~0.35 distance equivalence; empirical
        st['hits'] = min(10, st.get('hits', 0) + 1)
        if st['hits'] >= 3:
            recognized = True
    else:
        st['hits'] = 0

    # If recognized and not yet logged, log attendance once and redirect
    if recognized and not st.get('logged'):
        # Cooldown protection using cache
        cooldown_seconds = int(getattr(settings, 'ATTENDANCE_COOLDOWN_SECONDS', 90))
        cd_key = f'attendance_cooldown:{user.id}'
        if cache.get(cd_key):
            # Still in cooldown; return silent ok:false to keep loop running client-side
            label = 'COOLDOWN'
            head_color = (0, 200, 200)
            cv2.putText(bgr, label, (10, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.8, head_color, 2, cv2.LINE_AA)
            _, buffer = cv2.imencode('.jpg', bgr)
            processed_frame = base64.b64encode(buffer).decode('utf-8')
            return JsonResponse({'ok': False, 'reason': 'cooldown', 'frame': processed_frame})

        # Persist attendance log (same logic as finalize, without geofence)
        from .models import AttendanceLog
        now = timezone.localtime()
        today = timezone.localdate()
        # Block on public holidays (manual org + national)
        if is_holiday_for_org(cm.user, today).is_holiday:
            label = 'HOLIDAY'
            head_color = (0, 0, 255)
            cv2.putText(bgr, label, (10, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.8, head_color, 2, cv2.LINE_AA)
            _, buffer = cv2.imencode('.jpg', bgr)
            processed_frame = base64.b64encode(buffer).decode('utf-8')
            return JsonResponse({'ok': False, 'reason': 'holiday', 'frame': processed_frame})
        # Derive state from state_code if possible
        state = ''
        try:
            state = (cm.state_code or '').split('/')[0]
        except Exception:
            state = ''
        log, created = AttendanceLog.objects.get_or_create(
            account=user,
            date=today,
            defaults={
                'org': cm.user,
                'name': cm.full_name,
                'state': state,
                'code': cm.state_code or '',
            }
        )
        if created or not log.time_in:
            log.time_in = now.time()
        if not log.time_in or now.time() >= log.time_in:
            log.time_out = now.time()
        else:
            log.time_out = log.time_in
        log.save()
        st['logged'] = True
        cache.set(cd_key, True, timeout=max(1, cooldown_seconds))

        # Return JSON success with redirect path and message
        return JsonResponse({'ok': True, 'redirect': '/dashboard/', 'message': f'Attendance logged for {cm.full_name}'})

    # Otherwise, draw overlays and return 200 JSON (no redirect)
    label = 'RECOGNIZED' if recognized else 'Scanning…'
    head_color = (0, 200, 0) if recognized else (0, 200, 200)
    cv2.putText(bgr, label, (10, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.8, head_color, 2, cv2.LINE_AA)
    _, buffer = cv2.imencode('.jpg', bgr)
    processed_frame = base64.b64encode(buffer).decode('utf-8')
    # If no faces detected at all, hint for client
    reason = None
    if not recognized:
        if not face_locations:
            reason = 'no_face'
        else:
            reason = 'no_match'
    return JsonResponse({'ok': False, 'reason': reason or 'scanning', 'frame': processed_frame, 'recognized': recognized})


def attendance_finalize(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    user = request.user
    if getattr(user, 'role', None) != 'CORPER':
        return JsonResponse({'detail': 'Not allowed'}, status=403)
    cm = getattr(user, 'corper_profile', None)
    if not cm:
        return JsonResponse({'detail': 'Corper profile not found'}, status=404)
    st = _ATTENDANCE_STATE.get(cm.id, {'hits': 0})
    if st.get('hits', 0) < 3:
        return JsonResponse({'detail': 'Face not recognized'}, status=400)
    # Geofence check (require lat/lng in request)
    try:
        lat = float(request.POST.get('lat')) if request.method == 'POST' else None
        lng = float(request.POST.get('lng')) if request.method == 'POST' else None
    except Exception:
        lat = lng = None
    tgt_lat, tgt_lng, _ = _get_target_location_for_corper(cm)
    if tgt_lat is None or tgt_lng is None:
        return JsonResponse({'detail': 'Organization location not configured'}, status=400)
    if lat is None or lng is None:
        return JsonResponse({'detail': 'Missing location; enable location to mark attendance'}, status=400)
    threshold = getattr(settings, 'ATTENDANCE_GEOFENCE_METERS', 250)
    if _haversine_m(lat, lng, tgt_lat, tgt_lng) > threshold:
        return JsonResponse({'detail': 'You are not within the attendance proximity'}, status=403)
    # Block on public holidays (manual org + national)
    today = timezone.localdate()
    if is_holiday_for_org(cm.user, today).is_holiday:
        return JsonResponse({'detail': 'Today is a public holiday for your organization'}, status=403)
    # Persist attendance log: create/update today's record
    from .models import AttendanceLog
    now = timezone.localtime()
    # Derive state from state_code if possible
    state = ''
    try:
        state = (cm.state_code or '').split('/')[0]
    except Exception:
        state = ''
    log, created = AttendanceLog.objects.get_or_create(
        account=user,
        date=today,
        defaults={
            'org': cm.user,
            'name': cm.full_name,
            'state': state,
            'code': cm.state_code or '',
        }
    )
    # Set time_in if missing, and always refresh time_out to now (not earlier than time_in)
    if created or not log.time_in:
        log.time_in = now.time()
    # Always update time_out on check-in to reflect last seen time
    if not log.time_in or now.time() >= log.time_in:
        log.time_out = now.time()
    else:
        # Safety: never set time_out earlier than time_in
        log.time_out = log.time_in
    log.save()

    _reset_attendance_state(cm.id)
    return JsonResponse({'status': 'ok'})


class RegisterView(APIView):
    # AllowAny and disable SessionAuthentication/CSRF for this endpoint
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        serializer = OrganizationRegisterSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({
            'message': 'Registration successful. Please check your email to verify your account and set your password.'
        }, status=status.HTTP_201_CREATED)


class VerifyEmailView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        token = request.query_params.get('token')
        if not token:
            return Response({'detail': 'Missing token'}, status=status.HTTP_400_BAD_REQUEST)

        user_id = validate_email_token(token)
        if not user_id:
            return Response({'detail': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        user.is_active = True
        user.is_email_verified = True
        user.save(update_fields=['is_active', 'is_email_verified'])

        # If called from SPA (front=1), return JSON instead of redirecting
        if request.query_params.get('front') == '1':
            role = getattr(user, 'role', None)
            return Response({'verified': True, 'role': role})

        # Redirect to frontend page. If user has no usable password, include token for password setup
        # Prefer request Origin if it matches configured FRONTEND_ORIGINS for better DX (5173/5174)
        request_origin = request.headers.get('Origin')
        try:
            allowed = set(getattr(settings, 'FRONTEND_ORIGINS', []))
        except Exception:
            allowed = set()
        base = (request_origin if request_origin in allowed else getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173'))).rstrip('/')
        # If the user has no password (invited admin/corper), send to password set page with token
        role = request.query_params.get('role') or getattr(user, 'role', None)
        # For invited roles (branch admin, corper), always allow setting password on first verify
        if role in ('BRANCH', 'CORPER'):
            return redirect(f"{base}/verify-success?token={token}&role={role}")
        # For org accounts, only show password set if they don't already have one
        if not user.has_usable_password():
            return redirect(f"{base}/verify-success?token={token}&role={role or 'ORG'}")
        # Otherwise, show success page (no password needed)
        suffix = f"?role={role}" if role else ""
        return redirect(f"{base}/verify-success{suffix}")


class PasswordSetView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        token = request.data.get('token')
        password = request.data.get('password')
        password_confirm = request.data.get('password_confirm')
        if not token or not password:
            return Response({'detail': 'token and password are required'}, status=status.HTTP_400_BAD_REQUEST)
        if password_confirm is not None and password != password_confirm:
            return Response({'detail': 'Passwords do not match'}, status=status.HTTP_400_BAD_REQUEST)
        user_id = validate_email_token(token)
        if not user_id:
            return Response({'detail': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        user.set_password(password)
        user.is_active = True
        user.is_email_verified = True
        user.save(update_fields=['password', 'is_active', 'is_email_verified'])
        return Response({'message': 'Password set successfully'})


class PasswordResetRequestView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response({'detail': 'email is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            # Avoid user enumeration
            return Response({'message': 'If the email exists, a reset link has been sent.'})
        token = generate_email_token(user.id)
        base = getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')).rstrip('/')
        role = getattr(user, 'role', None)
        reset_url = f"{base}/reset-password?token={token}{f'&role={role}' if role else ''}"
        from django.core.mail import send_mail
        send_mail(
            'Reset your NYSC Clearance password',
            f"Use the link to reset your password:\n{reset_url}",
            settings.DEFAULT_FROM_EMAIL,
            [email],
        )
        return Response({'message': 'If the email exists, a reset link has been sent.'})


class PasswordResetConfirmView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        token = request.data.get('token')
        password = request.data.get('password')
        password_confirm = request.data.get('password_confirm')
        if not token or not password:
            return Response({'detail': 'token and password are required'}, status=status.HTTP_400_BAD_REQUEST)
        if password_confirm is not None and password != password_confirm:
            return Response({'detail': 'Passwords do not match'}, status=status.HTTP_400_BAD_REQUEST)
        user_id = validate_email_token(token)
        if not user_id:
            return Response({'detail': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'detail': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        user.set_password(password)
        user.save(update_fields=['password'])
        return Response({'message': 'Password reset successfully'})


class CSRFView(APIView):
    permission_classes = []
    authentication_classes = []

    def get(self, request):
        # Generate CSRF token, set cookie, and return it for debugging
        token = get_token(request)
        return Response({'csrfToken': token})


def _admin_config_version():
    timestamps = []
    for model_cls in (SystemSetting, SubscriptionPlanSetting, PaystackConfig):
        try:
            latest = model_cls.objects.aggregate(latest=models.Max('updated_at')).get('latest')
            if latest:
                timestamps.append(latest)
        except Exception:
            pass
    if not timestamps:
        return '0'
    latest = max(timestamps)
    return str(int(latest.timestamp() * 1000))


class ConfigVersionView(APIView):
    permission_classes = []
    authentication_classes = []

    def get(self, request):
        return Response({
            'version': _admin_config_version(),
            'deployment': getattr(settings, 'DEPLOYMENT_VERSION', '') or '',
            'checked_at': timezone.now().isoformat(),
        })


class ConfigView(APIView):
    """Expose non-sensitive runtime config for the frontend.

    Helps centralize CSRF/cookie names, base URLs, and origins across environments.
    """
    permission_classes = []
    authentication_classes = []

    def get(self, request):
        try:
            cors = list(getattr(settings, 'CORS_ALLOWED_ORIGINS', []) or [])
        except Exception:
            cors = []
        try:
            csrf_origins = list(getattr(settings, 'CSRF_TRUSTED_ORIGINS', []) or [])
        except Exception:
            csrf_origins = []
        data = {
            'api_base': getattr(settings, 'API_BASE_URL', ''),
            'frontend_base': getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', '')),
            'paystack_webhook_url': f"{getattr(settings, 'API_BASE_URL', '').rstrip('/')}/api/auth/paystack/webhook/",
            'csrf_cookie_name': getattr(settings, 'CSRF_COOKIE_NAME', 'csrftoken'),
            'session_cookie_name': getattr(settings, 'SESSION_COOKIE_NAME', 'sessionid'),
            'cors_allowed_origins': cors,
            'csrf_trusted_origins': csrf_origins,
            'cookie_same_site': {
                'session': getattr(settings, 'SESSION_COOKIE_SAMESITE', 'Lax'),
                'csrf': getattr(settings, 'CSRF_COOKIE_SAMESITE', 'Lax'),
            },
            'cookie_secure': {
                'session': bool(getattr(settings, 'SESSION_COOKIE_SECURE', False)),
                'csrf': bool(getattr(settings, 'CSRF_COOKIE_SECURE', False)),
            },
            'settings_version': _admin_config_version(),
            'debug': bool(getattr(settings, 'DEBUG', False)),
        }
        return Response(data)


class LoginView(APIView):
    # AllowAny and disable SessionAuthentication/CSRF for this endpoint
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        password = serializer.validated_data['password']
        expected_role = serializer.validated_data.get('role')
        user = authenticate(request, email=email, password=password)
        if not user:
            return Response({'detail': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)
        if expected_role and getattr(user, 'role', None) != expected_role:
            return Response({'detail': 'Invalid role for this login'}, status=status.HTTP_403_FORBIDDEN)
        if not user.is_active:
            return Response({'detail': 'Account not active'}, status=status.HTTP_403_FORBIDDEN)
        # Issue stateless token while keeping session login for backward compatibility
        try:
            from .auth import make_access_token
            token = make_access_token(user.id)
        except Exception:
            token = None
        # Optional: session login (kept for compatibility with existing clients)
        try:
            login(request, user)
        except Exception:
            pass
        OrganizationProfile.objects.get_or_create(user=user)
        resp = {'message': 'Logged in'}
        if token:
            resp['token'] = token
            resp['token_type'] = 'Bearer'
            resp['expires_in'] = 60 * 60 * 24
        return Response(resp)


class LogoutView(APIView):
    # Accept both session and token-based clients; no CSRF enforcement
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        logout(request)
        resp = Response({'message': 'Logged out'})
        # Explicitly expire cookies to prevent sticky auth across deployments/browsers.
        try:
            resp.delete_cookie(
                getattr(settings, 'SESSION_COOKIE_NAME', 'sessionid'),
                path=getattr(settings, 'SESSION_COOKIE_PATH', '/') or '/',
                domain=getattr(settings, 'SESSION_COOKIE_DOMAIN', None),
            )
        except Exception:
            pass
        try:
            resp.delete_cookie(
                getattr(settings, 'CSRF_COOKIE_NAME', 'csrftoken'),
                path=getattr(settings, 'CSRF_COOKIE_PATH', '/') or '/',
                domain=getattr(settings, 'CSRF_COOKIE_DOMAIN', None),
            )
        except Exception:
            pass
        return resp


class MeView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'authenticated': False})
        return Response({
            'authenticated': True,
            'email': request.user.email,
            'name': request.user.name,
            'role': getattr(request.user, 'role', 'ORG'),
        })


class ProfileView(APIView):
    def get(self, request):
        user = request.user
        org_user = user
        if getattr(user, 'role', None) == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if b:
                org_user = b.user
        elif getattr(user, 'role', None) == 'CORPER':
            try:
                org_user = user.corper_profile.user
            except Exception:
                pass
        profile, _ = OrganizationProfile.objects.get_or_create(user=org_user)
        return Response(OrganizationProfileSerializer(profile).data)

    def put(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organization can update profile')
        profile, _ = OrganizationProfile.objects.get_or_create(user=request.user)
        serializer = OrganizationProfileSerializer(profile, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


STRUCTURE_IMPORT_COLUMNS = [
    'branch_name',
    'branch_address',
    'branch_latitude',
    'branch_longitude',
    'admin_name',
    'admin_email',
    'admin_staff_id',
    'department_name',
    'unit_name',
]

CORPER_IMPORT_COLUMNS = [
    'full_name',
    'email',
    'gender',
    'state_code',
    'passing_out_date',
    'cds_day',
    'branch_name',
    'department_name',
    'unit_name',
]


def _clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _name_key(value):
    return re.sub(r'\s+', ' ', _clean_cell(value)).strip().lower()


def _truthy(value):
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'apply'}


def _read_import_rows(upload):
    if not upload:
        raise ValueError('Upload a CSV or Excel file.')

    filename = (getattr(upload, 'name', '') or '').lower()
    if filename.endswith('.csv'):
        text = upload.read().decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            raise ValueError('The file is empty or missing a header row.')
        rows = []
        for idx, row in enumerate(reader, start=2):
            cleaned = {str(k or '').strip().lower(): _clean_cell(v) for k, v in row.items()}
            if any(cleaned.values()):
                cleaned['_row'] = idx
                rows.append(cleaned)
        return rows

    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(upload.read()), read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(h or '').strip().lower() for h in (header_row or [])]
        if not any(headers):
            raise ValueError('The file is empty or missing a header row.')
        rows = []
        for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cleaned = {}
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                cleaned[header] = _clean_cell(values[col_idx] if col_idx < len(values) else '')
            if any(cleaned.values()):
                cleaned['_row'] = idx
                rows.append(cleaned)
        return rows

    raise ValueError('Unsupported file type. Upload .xlsx or .csv.')


def _read_xlsx_sheet_rows(wb, sheet_name):
    try:
        ws = wb[sheet_name]
    except Exception:
        return []
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(h or '').strip().lower() for h in (header_row or [])]
    if not any(headers):
        return []
    rows = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cleaned = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            cleaned[header] = _clean_cell(values[col_idx] if col_idx < len(values) else '')
        if any(cleaned.values()):
            cleaned['_row'] = idx
            cleaned['_sheet'] = sheet_name
            rows.append(cleaned)
    return rows


def _read_structure_import(upload):
    """Read structure import either as new multi-sheet workbook or legacy single-sheet rows."""
    filename = (getattr(upload, 'name', '') or '').lower()
    if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
        from openpyxl import load_workbook

        data = upload.read()
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheets = set(wb.sheetnames or [])
        multi = {'Branches', 'Departments', 'Units'}
        if multi.issubset(sheets):
            return {
                'mode': 'multi',
                'branches': _read_xlsx_sheet_rows(wb, 'Branches'),
                'departments': _read_xlsx_sheet_rows(wb, 'Departments'),
                'units': _read_xlsx_sheet_rows(wb, 'Units'),
            }
        # Legacy: read active sheet rows
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(h or '').strip().lower() for h in (header_row or [])]
        if not any(headers):
            raise ValueError('The file is empty or missing a header row.')
        rows = []
        for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cleaned = {}
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                cleaned[header] = _clean_cell(values[col_idx] if col_idx < len(values) else '')
            if any(cleaned.values()):
                cleaned['_row'] = idx
                cleaned['_sheet'] = ws.title
                rows.append(cleaned)
        return {'mode': 'legacy', 'rows': rows}

    # CSV is legacy combined rows
    return {'mode': 'legacy', 'rows': _read_import_rows(upload)}


def _xlsx_template_response(filename, sheet_name, columns, example_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    for row in example_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4F6228')
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 14), 32)

    # Optional dropdown for CDS day in corper template
    if 'cds_day' in columns:
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            idx = columns.index('cds_day') + 1
            col_letter = ws.cell(row=1, column=idx).column_letter
            dv = DataValidation(type="list", formula1='"Monday,Tuesday,Wednesday,Thursday,Friday"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}5000")
        except Exception:
            pass
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_workbook_response(filename, sheets):
    """Create a multi-sheet XLSX workbook response.

    sheets: list of dicts: { name, columns, rows }
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    # Remove default sheet; we'll add our own in order.
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F6228')

    for idx, spec in enumerate(sheets):
        ws = wb.create_sheet(spec['name'], idx)
        columns = spec.get('columns') or []
        ws.append(columns)
        for row in (spec.get('rows') or []):
            ws.append(row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 14), 32)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


class StructureImportTemplateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organization can download structure template')
        return _xlsx_workbook_response(
            'structure_import_template.xlsx',
            [
                {
                    'name': 'Branches',
                    'columns': [
                        'branch_name',
                        'branch_address',
                        'branch_latitude',
                        'branch_longitude',
                        'admin_name',
                        'admin_email',
                        'admin_staff_id',
                    ],
                    'rows': [
                        ['Head Office', '1 Main Road, Abuja', '9.0765', '7.3986', 'Amina Admin', 'admin@example.com', 'HQ-001'],
                        # Latitude/longitude are optional. You can update location later from the dashboard map.
                        ['Lagos Branch', '12 Marina, Lagos', '', '', '', '', ''],
                    ],
                },
                {
                    'name': 'Departments',
                    'columns': ['department_name'],
                    'rows': [
                        ['Human Resources'],
                        ['Finance'],
                    ],
                },
                {
                    'name': 'Units',
                    'columns': ['unit_name'],
                    'rows': [
                        ['Recruitment'],
                        ['Payroll'],
                    ],
                },
            ],
        )


class CorperImportTemplateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if getattr(request.user, 'role', None) not in {'ORG', 'BRANCH'}:
            raise PermissionDenied('Only organization or branch admins can download corper template')
        return _xlsx_template_response(
            'corpers_import_template.xlsx',
            'Corpers',
            CORPER_IMPORT_COLUMNS,
            [
                # department_name and unit_name are optional.
                ['Amina Yusuf', 'amina.corper@example.com', 'F', 'FC/24A/1234', '2026-10-31', 'Tuesday', 'Head Office', '', ''],
                ['David Okon', 'david.corper@example.com', 'M', 'LA/24B/5678', '2026-10-31', 'Thursday', 'Lagos Branch', 'Finance', 'Payroll'],
            ],
        )


class StructureImportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organization can import structure')
        upload = request.FILES.get('file')
        apply_changes = _truthy(request.data.get('apply'))
        try:
            data = _read_structure_import(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        preview = self._preview(request.user, data)
        if apply_changes and preview['errors_count']:
            return Response(preview, status=status.HTTP_400_BAD_REQUEST)
        if apply_changes:
            applied = self._apply(request, data)
            return Response({**self._preview(request.user, data), 'applied': applied})
        return Response(preview)

    def _preview(self, user, data):
        branches = { _name_key(b.name): b for b in BranchOffice.objects.filter(user=user) }
        departments = { _name_key(d.name): d for d in Department.objects.filter(user=user) }
        existing_units = { _name_key(u.name) for u in Unit.objects.filter(user=user) }
        file_branches = set()
        file_departments = set()
        file_units = set()
        preview_rows = []

        mode = data.get('mode')
        if mode == 'legacy':
            # Legacy combined rows: branch + department + unit in one sheet
            for row in (data.get('rows') or []):
                messages = []
                branch_name = row.get('branch_name', '')
                department_name = row.get('department_name', '')
                unit_name = row.get('unit_name', '')
                branch_key = _name_key(branch_name)
                department_key = _name_key(department_name)
                unit_key = _name_key(unit_name)
                if not branch_key:
                    messages.append('branch_name is required')
                if row.get('admin_email') and '@' not in row.get('admin_email'):
                    messages.append('admin_email is not valid')
                branch_exists = branch_key in branches or branch_key in file_branches
                department_exists = not department_key or department_key in departments or department_key in file_departments
                unit_exists = not unit_key or unit_key in existing_units or unit_key in file_units
                if branch_key:
                    file_branches.add(branch_key)
                if department_key:
                    file_departments.add(department_key)
                if unit_key:
                    file_units.add(unit_key)
                preview_rows.append({
                    'row': row.get('_row'),
                    'branch_name': branch_name,
                    'department_name': department_name,
                    'unit_name': unit_name,
                    'status': 'error' if messages else 'ok',
                    'messages': messages,
                    'branch_action': 'reuse' if branch_exists else 'create',
                    'department_action': '' if not department_key else ('reuse' if department_exists else 'create'),
                    'unit_action': '' if not unit_key else ('reuse' if unit_exists else 'create'),
                })
        else:
            # Multi-sheet import
            for row in (data.get('branches') or []):
                messages = []
                branch_name = row.get('branch_name', '')
                branch_key = _name_key(branch_name)
                if not branch_key:
                    messages.append('branch_name is required')
                if row.get('admin_email') and '@' not in row.get('admin_email'):
                    messages.append('admin_email is not valid')
                branch_exists = branch_key in branches or branch_key in file_branches
                if branch_key:
                    file_branches.add(branch_key)
                preview_rows.append({
                    'row': row.get('_row'),
                    'branch_name': branch_name,
                    'department_name': '',
                    'unit_name': '',
                    'status': 'error' if messages else 'ok',
                    'messages': messages,
                    'branch_action': 'reuse' if branch_exists else 'create',
                    'department_action': '',
                    'unit_action': '',
                })

            for row in (data.get('departments') or []):
                messages = []
                department_name = row.get('department_name', '') or row.get('name', '')
                department_key = _name_key(department_name)
                if not department_key:
                    messages.append('department_name is required')
                department_exists = department_key in departments or department_key in file_departments
                if department_key:
                    file_departments.add(department_key)
                preview_rows.append({
                    'row': row.get('_row'),
                    'branch_name': '',
                    'department_name': department_name,
                    'unit_name': '',
                    'status': 'error' if messages else 'ok',
                    'messages': messages,
                    'branch_action': '',
                    'department_action': 'reuse' if department_exists else 'create',
                    'unit_action': '',
                })

            for row in (data.get('units') or []):
                messages = []
                unit_name = row.get('unit_name', '')
                unit_key = _name_key(unit_name)
                if not unit_key:
                    messages.append('unit_name is required')
                if unit_key:
                    unit_exists = unit_key in existing_units or unit_key in file_units
                    file_units.add(unit_key)
                preview_rows.append({
                    'row': row.get('_row'),
                    'branch_name': '',
                    'department_name': '',
                    'unit_name': unit_name,
                    'status': 'error' if messages else 'ok',
                    'messages': messages,
                    'branch_action': '',
                    'department_action': '',
                    'unit_action': '' if not unit_key else ('reuse' if unit_exists else 'create'),
                })

        errors_count = sum(1 for row in preview_rows if row['status'] == 'error')
        return {
            'ok': errors_count == 0,
            'errors_count': errors_count,
            'summary': {
                'rows': len(preview_rows),
                'branches_to_create': sum(1 for key in file_branches if key not in branches),
                'departments_to_create': sum(1 for key in file_departments if key not in departments),
                'units_to_create': sum(1 for key in file_units if key not in existing_units),
            },
            'rows': preview_rows[:200],
        }

    @transaction.atomic
    def _apply(self, request, data):
        user = request.user
        created = {'branches': 0, 'departments': 0, 'units': 0}
        branches = { _name_key(b.name): b for b in BranchOffice.objects.select_for_update().filter(user=user) }
        departments = { _name_key(d.name): d for d in Department.objects.select_for_update().filter(user=user) }
        existing_units = { _name_key(u.name) for u in Unit.objects.select_for_update().filter(user=user) }

        if data.get('mode') == 'legacy':
            legacy_rows = data.get('rows') or []
            branch_rows = legacy_rows
            dept_rows = legacy_rows
            unit_rows = legacy_rows
        else:
            branch_rows = data.get('branches') or []
            dept_rows = data.get('departments') or []
            unit_rows = data.get('units') or []

        # Branches
        for row in branch_rows:
            branch_key = _name_key(row.get('branch_name'))
            if not branch_key:
                continue
            branch = branches.get(branch_key)
            if not branch:
                serializer = BranchOfficeSerializer(data={
                    'name': row.get('branch_name'),
                    'address': row.get('branch_address', ''),
                    'latitude': row.get('branch_latitude') or None,
                    'longitude': row.get('branch_longitude') or None,
                    'admin_name': row.get('admin_name', ''),
                    'admin_email': row.get('admin_email', ''),
                    'admin_staff_id': row.get('admin_staff_id', ''),
                }, context={'request': request})
                serializer.is_valid(raise_exception=True)
                branch = serializer.save()
                branches[branch_key] = branch
                created['branches'] += 1
            elif row.get('admin_email') and not branch.admin_id:
                serializer = BranchOfficeSerializer(branch, data={
                    'admin_name': row.get('admin_name', ''),
                    'admin_email': row.get('admin_email', ''),
                    'admin_staff_id': row.get('admin_staff_id', ''),
                }, partial=True, context={'request': request})
                serializer.is_valid(raise_exception=True)
                branch = serializer.save()
                branches[branch_key] = branch

        # Departments
        for row in dept_rows:
            dept_name = row.get('department_name') or row.get('name')
            department_key = _name_key(dept_name)
            if not department_key:
                continue
            if department_key not in departments:
                department = Department.objects.create(user=user, name=dept_name)
                departments[department_key] = department
                created['departments'] += 1

        # Units
        for row in unit_rows:
            unit_key = _name_key(row.get('unit_name'))
            if data.get('mode') == 'legacy':
                unit_key = _name_key(row.get('unit_name'))
            if not unit_key:
                continue
            if unit_key not in existing_units:
                Unit.objects.create(user=user, name=row.get('unit_name'))
                existing_units.add(unit_key)
                created['units'] += 1
        return created


class CorperImportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if getattr(request.user, 'role', None) not in {'ORG', 'BRANCH'}:
            raise PermissionDenied('Only organization or branch admins can import corpers')
        upload = request.FILES.get('file')
        apply_changes = _truthy(request.data.get('apply'))
        try:
            rows = _read_import_rows(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        preview = self._preview(request.user, rows)
        if apply_changes and preview['errors_count']:
            return Response(preview, status=status.HTTP_400_BAD_REQUEST)
        if apply_changes:
            applied = self._apply(request, preview['valid_payloads'])
            preview.pop('valid_payloads', None)
            return Response({**preview, 'applied': applied})
        preview.pop('valid_payloads', None)
        return Response(preview)

    def _scope(self, user):
        if getattr(user, 'role', None) == 'ORG':
            org_user = user
            branch_qs = BranchOffice.objects.filter(user=user)
            default_branch = None
        else:
            default_branch = BranchOffice.objects.filter(admin=user).first()
            org_user = default_branch.user if default_branch else None
            branch_qs = BranchOffice.objects.filter(id=getattr(default_branch, 'id', None))
        return org_user, default_branch, branch_qs

    def _preview(self, user, rows):
        org_user, default_branch, branch_qs = self._scope(user)
        branches = { _name_key(b.name): b for b in branch_qs }
        departments = { _name_key(d.name): d for d in Department.objects.filter(user=org_user) } if org_user else {}
        units = { _name_key(u.name): u for u in Unit.objects.filter(user=org_user) } if org_user else {}
        emails = {_name_key(r.get('email')) for r in rows if _name_key(r.get('email'))}
        state_codes = {_clean_cell(r.get('state_code')).upper() for r in rows if _clean_cell(r.get('state_code'))}
        existing_emails = {_name_key(email) for email in User.objects.filter(email__in=emails).values_list('email', flat=True)}
        existing_state_codes = set(CorpMember.objects.filter(user=org_user, state_code__in=state_codes).values_list('state_code', flat=True)) if org_user else set()
        seen_emails = set()
        seen_state_codes = set()
        preview_rows = []
        valid_payloads = []

        for row in rows:
            messages = []
            full_name = row.get('full_name', '')
            email = _name_key(row.get('email'))
            gender = _clean_cell(row.get('gender')).upper()
            gender_map = {'MALE': 'M', 'FEMALE': 'F', 'OTHER': 'O'}
            gender = gender_map.get(gender, gender)
            state_code = _clean_cell(row.get('state_code')).upper()
            branch_key = _name_key(row.get('branch_name')) if getattr(user, 'role', None) == 'ORG' else _name_key(getattr(default_branch, 'name', ''))
            department_key = _name_key(row.get('department_name'))
            unit_key = _name_key(row.get('unit_name'))

            if not full_name:
                messages.append('full_name is required')
            if not email:
                messages.append('email is required')
            elif email in seen_emails:
                messages.append('email is duplicated in this file')
            elif email in existing_emails:
                messages.append('email already exists')
            if gender not in {'M', 'F', 'O'}:
                messages.append('gender must be M, F, O, Male, Female, or Other')
            if not re.match(r'^[A-Z]{2}/\d{2}[A-Z]/\d{4}$', state_code or ''):
                messages.append('state_code must be in format AA/00A/0000')
            elif state_code in seen_state_codes:
                messages.append('state_code is duplicated in this file')
            elif state_code in existing_state_codes:
                messages.append('state_code already exists in this organization')
            if not row.get('passing_out_date'):
                messages.append('passing_out_date is required')
            if not branch_key:
                messages.append('branch_name is required')
            branch = branches.get(branch_key)
            if branch_key and not branch:
                messages.append('branch_name was not found')
            department = departments.get(department_key) if department_key else None
            if department_key and not department:
                messages.append('department_name was not found in this organisation')
            unit = units.get(unit_key) if unit_key else None
            if unit_key and not unit:
                messages.append('unit_name was not found in this organisation')
            cds_day = _clean_cell(row.get('cds_day', ''))
            cds_int = None
            if cds_day != '':
                day_map = {'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3, 'friday': 4}
                key = cds_day.strip().lower()
                if key in day_map:
                    cds_int = day_map[key]
                else:
                    try:
                        cds_int = int(float(cds_day))
                        if cds_int < 0 or cds_int > 4:
                            messages.append('cds_day must be Monday-Friday or 0-4')
                    except Exception:
                        messages.append('cds_day must be Monday-Friday or 0-4')
                        cds_int = None

            if email:
                seen_emails.add(email)
            if state_code:
                seen_state_codes.add(state_code)

            payload = {
                'full_name': full_name,
                'email': email,
                'gender': gender,
                'state_code': state_code,
                'passing_out_date': row.get('passing_out_date'),
                'cds_day': cds_int,
                'branch': getattr(branch, 'id', None),
                'department': getattr(department, 'id', None) if department else None,
                'unit': getattr(unit, 'id', None) if unit else None,
            }
            if not messages:
                valid_payloads.append(payload)
            preview_rows.append({
                'row': row.get('_row'),
                'full_name': full_name,
                'email': email,
                'state_code': state_code,
                'branch_name': row.get('branch_name') or getattr(default_branch, 'name', ''),
                'department_name': row.get('department_name', ''),
                'unit_name': row.get('unit_name', ''),
                'status': 'error' if messages else 'ok',
                'messages': messages,
            })

        errors_count = sum(1 for row in preview_rows if row['status'] == 'error')
        return {
            'ok': errors_count == 0,
            'errors_count': errors_count,
            'summary': {
                'rows': len(rows),
                'corpers_to_create': len(valid_payloads),
                'face_capture': 'Face capture remains live after import.',
            },
            'rows': preview_rows[:200],
            'valid_payloads': valid_payloads,
        }

    @transaction.atomic
    def _apply(self, request, payloads):
        created = 0
        for payload in payloads:
            serializer = CorpMemberSerializer(data=payload, context={'request': request})
            serializer.is_valid(raise_exception=True)
            serializer.save()
            created += 1
        return {'corpers': created}


class BranchOfficeViewSet(viewsets.ModelViewSet):
    serializer_class = BranchOfficeSerializer

    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'role', None) == 'ORG':
            return BranchOffice.objects.filter(user=user)
        if getattr(user, 'role', None) == 'BRANCH':
            return BranchOffice.objects.filter(admin=user)
        if getattr(user, 'role', None) == 'CORPER':
            try:
                br_id = user.corper_profile.branch_id
                return BranchOffice.objects.filter(id=br_id)
            except Exception:
                return BranchOffice.objects.none()
        return BranchOffice.objects.none()

    def perform_create(self, serializer):
        # user is set inside the serializer using request from context
        serializer.save()

    @action(detail=True, methods=['post'])
    def clone_structure(self, request, pk=None):
        """No-op for org-wide departments/units.

        Body: { "source": <branch_id> }
        """
        return Response({'detail': 'clone_structure is not required for organisation-wide departments and units.'})


class DepartmentViewSet(viewsets.ModelViewSet):
    serializer_class = DepartmentSerializer
    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'role', None) == 'ORG':
            return Department.objects.filter(user=user)
        if getattr(user, 'role', None) == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            org_user = b.user if b else None
            return Department.objects.filter(user=org_user) if org_user else Department.objects.none()
        if getattr(user, 'role', None) == 'CORPER':
            try:
                org_user = user.corper_profile.user
                return Department.objects.filter(user=org_user)
            except Exception:
                return Department.objects.none()
        return Department.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if getattr(user, 'role', None) not in {'ORG', 'BRANCH'}:
            raise PermissionDenied('Not allowed')
        serializer.save()


class UnitViewSet(viewsets.ModelViewSet):
    serializer_class = UnitSerializer
    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'role', None) == 'ORG':
            return Unit.objects.filter(user=user)
        if getattr(user, 'role', None) == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            org_user = b.user if b else None
            return Unit.objects.filter(user=org_user) if org_user else Unit.objects.none()
        if getattr(user, 'role', None) == 'CORPER':
            try:
                org_user = user.corper_profile.user
                return Unit.objects.filter(user=org_user)
            except Exception:
                return Unit.objects.none()
        return Unit.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if getattr(user, 'role', None) == 'ORG':
            serializer.save()
            return
        elif getattr(user, 'role', None) == 'BRANCH':
            serializer.save()
            return
        else:
            raise PermissionDenied('Not allowed')


class CorpMemberViewSet(viewsets.ModelViewSet):
    serializer_class = CorpMemberSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'ORG':
            # Include any corpers owned by this org or whose branch belongs to this org
            return CorpMember.objects.filter(models.Q(user=user) | models.Q(branch__user=user)).distinct()
        if user.role == 'BRANCH':
            branches = BranchOffice.objects.filter(admin=user)
            return CorpMember.objects.filter(branch__in=branches)
        if user.role == 'CORPER':
            try:
                return CorpMember.objects.filter(account=user)
            except Exception:
                return CorpMember.objects.none()
        return CorpMember.objects.none()

    def perform_create(self, serializer):
        # Serializer will map the owning organization correctly and default branch for branch admins
        serializer.save()

    def perform_update(self, serializer):
        """Restrict branch admins to editing only their corpers and to their own branches/departments/units."""
        user = self.request.user
        instance = self.get_object()
        if getattr(user, 'role', None) == 'BRANCH':
            # Ensure the corper is in one of the admin's branches
            admin_branch = BranchOffice.objects.filter(admin=user)
            if not admin_branch.filter(id=getattr(instance.branch, 'id', None)).exists():
                raise PermissionDenied('Not allowed to edit this corper')
            # If updating branch/department/unit, constrain them to admin's branches
            data = serializer.validated_data
            new_branch = data.get('branch') or instance.branch
            if new_branch and not admin_branch.filter(id=new_branch.id).exists():
                raise PermissionDenied('Cannot move corper outside your branch')
            # Validate department/unit belong to the organisation (not nested, not office-scoped)
            dept = data.get('department')
            unit = data.get('unit')
            org_user = instance.user
            if dept and getattr(dept, 'user_id', None) != getattr(org_user, 'id', None):
                raise PermissionDenied('Invalid department for this organisation')
            if unit and getattr(unit, 'user_id', None) != getattr(org_user, 'id', None):
                raise PermissionDenied('Invalid unit for this organisation')
        serializer.save()


class PublicHolidayViewSet(viewsets.ModelViewSet):
    serializer_class = PublicHolidaySerializer

    def get_queryset(self):
        user = self.request.user
        if getattr(user, 'role', 'ORG') == 'ORG':
            return PublicHoliday.objects.filter(user=user).order_by('start_date')
        # For branch admin/corper, map to their org
        org_user = None
        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            org_user = b.user if b else None
        elif user.role == 'CORPER':
            try:
                org_user = user.corper_profile.user
            except Exception:
                org_user = None
        return PublicHoliday.objects.filter(user=org_user).order_by('start_date')

    def perform_create(self, serializer):
        if self.request.user.role != 'ORG':
            raise PermissionDenied('Only organization can create holidays')
        serializer.save(user=self.request.user)


class AllHolidaysView(APIView):
    """Return combined national + organisation manual holidays.

    National holidays are read-only and apply to all orgs.
    Manual holidays are organization-scoped and deletable.
    """

    def get(self, request):
        user = request.user
        if not getattr(user, 'authenticated', True):
            return Response({'detail': 'Not authenticated'}, status=401)

        # Determine org user
        org_user = user if user.role == 'ORG' else None
        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            org_user = b.user if b else None
        elif user.role == 'CORPER':
            try:
                org_user = user.corper_profile.user
            except Exception:
                org_user = None

        year = timezone.localdate().year
        # Ensure we have national holidays (lazy auto-sync)
        ensure_national_holidays(year=year, country_code='NG')
        ensure_national_holidays(year=year + 1, country_code='NG')
        # National holidays for current year (and next year for smoother UX)
        national = NationalHoliday.objects.filter(country_code='NG', date__year__in=[year, year + 1]).order_by('date')
        manual = PublicHoliday.objects.filter(user=org_user).order_by('start_date') if org_user else PublicHoliday.objects.none()

        out = []
        for h in national:
            out.append({
                'id': f'national:{h.id}',
                'source': 'NATIONAL',
                'title': h.name,
                'start_date': h.date.isoformat(),
                'end_date': h.date.isoformat(),
                'deletable': False,
            })
        for h in manual:
            out.append({
                'id': h.id,
                'source': 'MANUAL',
                'title': h.title,
                'start_date': h.start_date.isoformat() if h.start_date else None,
                'end_date': h.end_date.isoformat() if h.end_date else None,
                'deletable': user.role == 'ORG',
            })
        return Response(out)


class LeaveRequestViewSet(viewsets.ModelViewSet):
    serializer_class = LeaveRequestSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'ORG':
            return LeaveRequest.objects.filter(corper__user=user).order_by('-created_at')
        if user.role == 'BRANCH':
            branches = BranchOffice.objects.filter(admin=user)
            return LeaveRequest.objects.filter(branch__in=branches).order_by('-created_at')
        if user.role == 'CORPER':
            try:
                cm = user.corper_profile
            except Exception:
                return LeaveRequest.objects.none()
            return LeaveRequest.objects.filter(corper=cm).order_by('-created_at')
        return LeaveRequest.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != 'CORPER':
            raise PermissionDenied('Only corpers can create leave requests')
        cm = getattr(user, 'corper_profile', None)
        if not cm:
            raise PermissionDenied('Corper profile not found')
        serializer.save(corper=cm, branch=cm.branch)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        lr = self.get_object()
        user = request.user
        if user.role not in ('ORG','BRANCH'):
            raise PermissionDenied('Not allowed')
        if user.role == 'BRANCH' and lr.branch and lr.branch.admin_id != user.id:
            raise PermissionDenied('Not your branch')
        lr.status = 'APPROVED'
        lr.decided_by = user
        lr.save(update_fields=['status','decided_by','updated_at'])
        return Response({'status': 'APPROVED'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        lr = self.get_object()
        user = request.user
        if user.role not in ('ORG','BRANCH'):
            raise PermissionDenied('Not allowed')
        if user.role == 'BRANCH' and lr.branch and lr.branch.admin_id != user.id:
            raise PermissionDenied('Not your branch')
        lr.status = 'REJECTED'
        lr.decided_by = user
        lr.save(update_fields=['status','decided_by','updated_at'])
        return Response({'status': 'REJECTED'})


class StatsView(APIView):
    def get(self, request):
        user = request.user
        # Scope stats by role: ORG sees org-wide, BRANCH sees their managed branch only
        if getattr(user, 'role', None) == 'BRANCH':
            branch_qs = BranchOffice.objects.filter(admin=user)
            corpers_qs = CorpMember.objects.filter(branch__in=branch_qs)
            org_user = branch_qs.first().user if branch_qs.exists() else None
            departments_qs = Department.objects.filter(user=org_user) if org_user else Department.objects.none()
            units_qs = Unit.objects.filter(user=org_user) if org_user else Unit.objects.none()
            # Attendance logs for accounts under these branches
            att_qs = AttendanceLog.objects.filter(account__corper_profile__branch__in=branch_qs)
        else:
            branch_qs = BranchOffice.objects.filter(user=user)
            corpers_qs = CorpMember.objects.filter(user=user)
            departments_qs = Department.objects.filter(user=user)
            units_qs = Unit.objects.filter(user=user)
            # Attendance logs for this organization
            if getattr(user, 'role', None) == 'CORPER':
                att_qs = AttendanceLog.objects.filter(account=user)
            else:
                att_qs = AttendanceLog.objects.filter(org=user)

        by_branch_qs = (
            corpers_qs.values('branch__name')
                      .annotate(count=Count('id'))
                      .order_by('branch__name')
        )
        corpers_by_branch = []
        for row in by_branch_qs:
            name = row['branch__name'] or 'Unassigned'
            corpers_by_branch.append({'branch': name, 'count': row['count']})

        data = {
            'totals': {
                'branches': branch_qs.count(),
                'departments': departments_qs.count(),
                'units': units_qs.count(),
                'corpers': corpers_qs.count(),
            },
            'corpers_by_branch': corpers_by_branch,
            'attendance': _attendance_stats(att_qs)
        }
        return Response(data)

def _attendance_stats(att_qs):
    """Return counts for today, this month, and last 7 days timeline.

    Also includes per-day hours (float) computed where both time_in and time_out are present.
    For org/branch scopes (multiple accounts), hours are summed per date.
    """
    today = timezone.localdate()
    start_month = today.replace(day=1)
    today_qs = att_qs.filter(date=today)
    today_count = today_qs.count()
    month_qs = att_qs.filter(date__gte=start_month, date__lte=today)
    month_count = month_qs.count()
    last7 = []
    for i in range(6, -1, -1):
        d = today - timezone.timedelta(days=i)
        day_qs = att_qs.filter(date=d)
        c = day_qs.count()
        # Sum hours across logs for the date
        hours = 0.0
        for log in day_qs:
            if log.time_in and log.time_out:
                from datetime import datetime
                try:
                    start_dt = datetime.combine(d, log.time_in)
                    end_dt = datetime.combine(d, log.time_out)
                    delta = end_dt - start_dt
                    if delta.total_seconds() > 0:
                        hours += round(delta.total_seconds() / 3600.0, 2)
                except Exception:
                    pass
        last7.append({'date': d.isoformat(), 'count': c, 'hours': round(hours, 2)})
    return {
        'today': today_count,
        'this_month': month_count,
        'last7': last7,
    }


def _prev_month_bounds(today=None):
    today = today or timezone.localdate()
    first_this = today.replace(day=1)
    last_prev = first_this - timezone.timedelta(days=1)
    first_prev = last_prev.replace(day=1)
    return first_prev, last_prev


def _working_days(user, start_date, end_date):
    """Return list of working dates excluding weekends and holidays."""
    return working_days(user, start_date, end_date)


def _working_days_for_corper(cm, start_date, end_date):
    """Working days for a specific corper.

    Excludes weekends, holidays, and the corper's CDS day (if set).
    """

    exclude = getattr(cm, 'cds_day', None)
    try:
        exclude = int(exclude) if exclude is not None else None
    except Exception:
        exclude = None
    if exclude is not None and (exclude < 0 or exclude > 4):
        exclude = None
    return working_days(cm.user, start_date, end_date, exclude_weekday=exclude)


def verify_clearance(request):
    """Public endpoint to verify a clearance reference.

    Expected format: NYSC-<STATE_CODE>-<YYYYMM>
    Example: NYSC-FC/23C/2354-202510
    """
    ref = request.GET.get('ref', '').strip()
    ctx = { 'ref': ref, 'valid': False, 'error': None }
    try:
        if not ref.startswith('NYSC-'):
            ctx['error'] = 'Invalid reference format'
            return render(request, 'verify_clearance.html', ctx, status=400)
        rest = ref[5:]
        dash = rest.rfind('-')
        if dash == -1:
            ctx['error'] = 'Invalid reference format'
            return render(request, 'verify_clearance.html', ctx, status=400)
        state_code = rest[:dash]
        yyyymm = rest[dash+1:]
        if len(yyyymm) != 6 or not yyyymm.isdigit():
            ctx['error'] = 'Invalid month segment'
            return render(request, 'verify_clearance.html', ctx, status=400)
        year = int(yyyymm[:4])
        month = int(yyyymm[4:])
        if month < 1 or month > 12:
            ctx['error'] = 'Invalid month value'
            return render(request, 'verify_clearance.html', ctx, status=400)

        # Locate the corper by state code
        cm = CorpMember.objects.filter(state_code=state_code).select_related('user', 'account', 'branch').first()
        if not cm:
            ctx['error'] = 'Reference not found'
            return render(request, 'verify_clearance.html', ctx, status=404)

        # Compute the month bounds for provided year-month
        from datetime import date
        import calendar
        start = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end = date(year, month, last_day)

        # Basic corroboration: ensure attendance logs exist in that month for the account (optional)
        has_any_log = AttendanceLog.objects.filter(account=cm.account, date__gte=start, date__lte=end).exists()

        # Organization details / signatory
        prof = OrganizationProfile.objects.filter(user=cm.user).first()
        signatory_name = getattr(prof, 'signatory_name', '') if prof else ''
        signature_url = getattr(prof.signature, 'url', '') if prof and getattr(prof, 'signature', None) else ''

        ctx.update({
            'valid': True,
            'state_code': cm.state_code,
            'name': cm.full_name,
            'org_name': getattr(cm.user, 'name', ''),
            'year': year,
            'month': start.strftime('%B'),
            'has_any_log': has_any_log,
            'signatory_name': signatory_name,
            'signature_url': signature_url,
        })
        return render(request, 'verify_clearance.html', ctx)
    except Exception:
        ctx['error'] = 'An unexpected error occurred'
        return render(request, 'verify_clearance.html', ctx, status=500)


def performance_summary(request):
    if getattr(request.user, 'role', None) != 'CORPER':
        raise PermissionDenied('Only corpers can access')
    cm = getattr(request.user, 'corper_profile', None)
    if not cm:
        return JsonResponse({'detail': 'Corper profile not found'}, status=404)
    start, end = _prev_month_bounds()
    work_days = _working_days(cm.user, start, end)
    work_set = set(work_days)
    logs = AttendanceLog.objects.filter(account=request.user, date__gte=start, date__lte=end)

    present_dates = set(logs.values_list('date', flat=True))
    present = len(present_dates & work_set)

    # Late threshold: org profile late_time; if missing, consider none late
    prof = OrganizationProfile.objects.filter(user=cm.user).first()
    late_time = getattr(prof, 'late_time', None)
    late = 0
    if late_time:
        for log in logs:
            if log.date in work_set and log.time_in and log.time_in > late_time:
                late += 1
    absent = max(0, len(work_days) - present)
    on_time = max(0, present - late)

    data = {
        'month': start.strftime('%B %Y').upper(),
        'range': {'start': start.isoformat(), 'end': end.isoformat()},
        'working_days': len(work_days),
        'present': present,
        'absent': absent,
        'late': late,
        'on_time': on_time,
        'name': cm.full_name.upper(),
        'state_code': cm.state_code,
    }
    return JsonResponse(data)


def performance_clearance_page(request):
    if getattr(request.user, 'role', None) != 'CORPER':
        raise PermissionDenied('Only corpers can access')
    cm = getattr(request.user, 'corper_profile', None)
    if not cm:
        return HttpResponseNotFound('Corper profile not found')
    start, end = _prev_month_bounds()
    # Organization logo
    prof = OrganizationProfile.objects.filter(user=cm.user).first()
    logo_url = getattr(prof.logo, 'url', '') if prof and getattr(prof, 'logo', None) else ''
    # Signatory details from organization profile
    signature_url = getattr(prof.signature, 'url', '') if prof and getattr(prof, 'signature', None) else ''

    # Pronoun based on gender (possessive)
    gender = (cm.gender or '').upper()
    if gender == 'M':
        pronoun = 'his'
    elif gender == 'F':
        pronoun = 'her'
    else:
        pronoun = 'their'

    # Organization contact details
    org_address = getattr(cm.user, 'address', '') or ''
    org_email = getattr(cm.user, 'email', '') or ''
    org_phone = getattr(cm.user, 'phone_number', '') or ''

    ref_number = f"NYSC-{cm.state_code}-{start.strftime('%Y%m')}"
    verification_url = request.build_absolute_uri(f'/verify/?ref={ref_number}')

    # Eligibility check: previous month lateness/absence vs org thresholds
    # Skip this check if it's the corper's first clearance (no prior clearance debits found)
    is_first_clearance = (
        not _clearance_access_exists(cm) and
        not WalletTransaction.objects.filter(type='DEBIT', reference__startswith=f"NYSC-{cm.state_code}-").exists()
    )
    # Also skip penalty if the corper was enrolled after the clearance month started.
    enrolled_after_start = False
    try:
        joined = getattr(request.user, 'date_joined', None)
        enrolled_after_start = bool(joined and joined.date() > start)
    except Exception:
        enrolled_after_start = False
    # Allow if override exists for this month
    yyyymm = start.strftime('%Y%m')
    has_override = ClearanceOverride.objects.filter(corper=cm, year_month=yyyymm).exists()
    if not is_first_clearance and not has_override and not enrolled_after_start:
        work_days = _working_days_for_corper(cm, start, end)
        work_set = set(work_days)
        logs = AttendanceLog.objects.filter(account=request.user, date__gte=start, date__lte=end)
        all_present_dates = set(logs.values_list('date', flat=True))
        present_required = len(all_present_dates & work_set)
        cds_day = getattr(cm, 'cds_day', None)
        present_cds = 0
        try:
            cds_int = int(cds_day) if cds_day is not None else None
        except Exception:
            cds_int = None
        if cds_int is not None and 0 <= cds_int <= 4:
            present_cds = len([d for d in all_present_dates if d.weekday() == cds_int])
        present = min(len(work_days), present_required + present_cds)
        late_time = getattr(prof, 'late_time', None)
        late = 0
        if late_time:
            for log in logs:
                if log.date in work_set and log.time_in and log.time_in > late_time:
                    late += 1
        absent = max(0, len(work_days) - present)
        exceeded_absent = (
            getattr(prof, 'max_days_absent', None) is not None and absent > (prof.max_days_absent or 0)
        )
        exceeded_late = (
            getattr(prof, 'max_days_late', None) is not None and late > (prof.max_days_late or 0)
        )
        if exceeded_absent or exceeded_late:
            # Compute frontend base for links
            request_origin = request.headers.get('Origin')
            try:
                allowed = set(getattr(settings, 'FRONTEND_ORIGINS', []))
            except Exception:
                allowed = set()
            frontend_base = (request_origin if request_origin in allowed else getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173'))).rstrip('/')

            return render(request, 'clearance_restriction.html', {
                'month': start.strftime('%B %Y'),
                'absent': absent,
                'late': late,
                'max_absent': getattr(prof, 'max_days_absent', None),
                'max_late': getattr(prof, 'max_days_late', None),
                'contact_url': f'{frontend_base}/dashboard',
            }, status=403)

    # Authorize clearance once per corper per month:
    # subscription coverage first, then org wallet, admin wallet, and corper wallet.
    try:
        charged, charge_reason, charge_source = _authorize_clearance_access(
            cm,
            request.user,
            ref_number,
            debit_label='Clearance view charge',
        )
    except Exception:
        charged = False
        charge_reason = 'We could not complete the clearance payment check. Please try again or contact support.'
        charge_source = None

    if not charged:
        # Deny access nicely with a prompt to fund wallet
        # Compute frontend base for links
        request_origin = request.headers.get('Origin')
        try:
            allowed = set(getattr(settings, 'FRONTEND_ORIGINS', []))
        except Exception:
            allowed = set()
        frontend_base = (request_origin if request_origin in allowed else getattr(settings, 'FRONTEND_ORIGIN', getattr(settings, 'FRONTEND_URL', 'http://localhost:5173'))).rstrip('/')

        return render(request, 'clearance_payment_required.html', {
            'reason': charge_reason,
            'next_steps': 'Please ask your organization to renew or upgrade its subscription, or fund the organization, admin, or corper wallet and try again.',
            'fund_url': f'{frontend_base}/dashboard?fund=1',
            'dashboard_url': f'{frontend_base}/dashboard',
        }, status=402)

    # Map NYSC two-letter state code to state/capital for template placeholders
    STATE_MAP = {
        'AB': {'state': 'Abia', 'capital': 'Umuahia'},
        'AD': {'state': 'Adamawa', 'capital': 'Yola'},
        'AK': {'state': 'Akwa Ibom', 'capital': 'Uyo'},
        'AN': {'state': 'Anambra', 'capital': 'Awka'},
        'BA': {'state': 'Bauchi', 'capital': 'Bauchi'},
        'BY': {'state': 'Bayelsa', 'capital': 'Yenagoa'},
        'BN': {'state': 'Benue', 'capital': 'Makurdi'},
        'BO': {'state': 'Borno', 'capital': 'Maiduguri'},
        'CR': {'state': 'Cross River', 'capital': 'Calabar'},
        'DT': {'state': 'Delta', 'capital': 'Asaba'},
        'EB': {'state': 'Ebonyi', 'capital': 'Abakaliki'},
        'ED': {'state': 'Edo', 'capital': 'Benin City'},
        'EK': {'state': 'Ekiti', 'capital': 'Ado-Ekiti'},
        'EN': {'state': 'Enugu', 'capital': 'Enugu'},
        'FC': {'state': 'FCT', 'capital': 'Abuja'},
        'GM': {'state': 'Gombe', 'capital': 'Gombe'},
        'IM': {'state': 'Imo', 'capital': 'Owerri'},
        'JG': {'state': 'Jigawa', 'capital': 'Dutse'},
        'KD': {'state': 'Kaduna', 'capital': 'Kaduna'},
        'KN': {'state': 'Kano', 'capital': 'Kano'},
        'KT': {'state': 'Katsina', 'capital': 'Katsina'},
        'KB': {'state': 'Kebbi', 'capital': 'Birnin Kebbi'},
        'KG': {'state': 'Kogi', 'capital': 'Lokoja'},
        'KW': {'state': 'Kwara', 'capital': 'Ilorin'},
        'LA': {'state': 'Lagos', 'capital': 'Ikeja'},
        'NS': {'state': 'Nasarawa', 'capital': 'Lafia'},
        'NG': {'state': 'Niger', 'capital': 'Minna'},
        'OG': {'state': 'Ogun', 'capital': 'Abeokuta'},
        'OD': {'state': 'Ondo', 'capital': 'Akure'},
        'OS': {'state': 'Osun', 'capital': 'Osogbo'},
        'OY': {'state': 'Oyo', 'capital': 'Ibadan'},
        'PL': {'state': 'Plateau', 'capital': 'Jos'},
        'RV': {'state': 'Rivers', 'capital': 'Port Harcourt'},
        'SO': {'state': 'Sokoto', 'capital': 'Sokoto'},
        'TR': {'state': 'Taraba', 'capital': 'Jalingo'},
        'YB': {'state': 'Yobe', 'capital': 'Damaturu'},
        'ZM': {'state': 'Zamfara', 'capital': 'Gusau'},
    }

    raw_code = (cm.state_code or '').upper()
    # Extract first two letters as state code (robust to formats like LA/20B/1234)
    code_letters = ''.join([ch for ch in raw_code if ch.isalpha()])[:2]
    region = STATE_MAP.get(code_letters, {'state': '', 'capital': ''})

    ctx = {
        'reference_number': ref_number,
        'date': timezone.localdate().strftime('%Y-%m-%d'),
        'month': start.strftime('%B').upper(),
        'year': start.strftime('%Y'),
        'name': cm.full_name.upper(),
        'state_code': cm.state_code,
        'state': region['state'],
        'capital': region['capital'],
        'signatory_name': (getattr(prof, 'signatory_name', '') if prof else ''),
        'pronoun': pronoun,
        'logo_url': logo_url,
        'signature_url': signature_url,
        'verification_url': verification_url,
        'org_address': org_address,
        'org_email': org_email,
        'org_phone': org_phone,
    }
    return render(request, 'performance_clearance.html', ctx)

    def _attendance_stats(self, att_qs):
        """Return counts for today, this month, and last 7 days timeline."""
        today = timezone.localdate()
        start_month = today.replace(day=1)
        # Today count: any log present (time_in set counts as present)
        today_count = att_qs.filter(date=today).count()
        month_count = att_qs.filter(date__gte=start_month, date__lte=today).count()
        # Last 7 days (inclusive today)
        last7 = []
        for i in range(6, -1, -1):
            d = today - timezone.timedelta(days=i)
            c = att_qs.filter(date=d).count()
            last7.append({'date': d.isoformat(), 'count': c})
        return {
            'today': today_count,
            'this_month': month_count,
            'last7': last7,
        }


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'ORG':
            return Notification.objects.filter(user=user).order_by('-created_at')
        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            org = b.user if b else None
            return Notification.objects.filter(user=org).order_by('-created_at')
        if user.role == 'CORPER':
            try:
                cm = user.corper_profile
            except Exception:
                return Notification.objects.none()
            return Notification.objects.filter(user=cm.user).filter(models.Q(branch__isnull=True) | models.Q(branch=cm.branch)).order_by('-created_at')
        return Notification.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'ORG':
            serializer.save(user=user, created_by=user)
            return
        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if not b:
                raise PermissionDenied('No managed branch')
            serializer.save(user=b.user, branch=b, created_by=user)
            return
        raise PermissionDenied('Not allowed')


class QueryRecordViewSet(viewsets.ModelViewSet):
    serializer_class = QueryRecordSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'ORG':
            return QueryRecord.objects.filter(org=user).select_related('branch', 'corper', 'created_by').order_by('-created_at')
        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if not b:
                return QueryRecord.objects.none()
            return QueryRecord.objects.filter(org=b.user, branch=b).select_related('branch', 'corper', 'created_by').order_by('-created_at')
        if user.role == 'CORPER':
            try:
                cm = user.corper_profile
            except Exception:
                return QueryRecord.objects.none()
            return QueryRecord.objects.filter(corper=cm).select_related('branch', 'corper', 'created_by').order_by('-created_at')
        return QueryRecord.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        corper = serializer.validated_data.get('corper')
        if not corper:
            raise PermissionDenied('Corper is required')

        if user.role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if not b:
                raise PermissionDenied('No branch assigned')
            if corper.branch_id != b.id:
                raise PermissionDenied('Corper does not belong to your branch')
            serializer.save(org=b.user, branch=b, created_by=user)
            return

        # ORG
        if corper.user_id != user.id:
            raise PermissionDenied('Corper does not belong to your organization')
        serializer.save(org=user, branch=corper.branch, created_by=user)

    @action(detail=True, methods=['post'])
    def reply(self, request, pk=None):
        obj = self.get_object()
        if request.user.role != 'CORPER':
            raise PermissionDenied('Not allowed')
        try:
            cm = request.user.corper_profile
        except Exception:
            raise PermissionDenied('No corper profile')
        if obj.corper_id != cm.id:
            raise PermissionDenied('Not allowed')

        payload = request.data or {}
        reply_text = (payload.get('reply') or payload.get('message') or '').strip()
        if not reply_text:
            return Response({'detail': 'reply is required'}, status=400)

        obj.corper_reply = reply_text
        obj.replied_at = timezone.now()
        obj.replied_by = request.user
        obj.save(update_fields=['corper_reply', 'replied_at', 'replied_by', 'updated_at'])
        return Response({'status': 'replied'})

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        obj = self.get_object()
        if request.user.role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')
        obj.status = 'RESOLVED'
        obj.resolved_by = request.user
        obj.save(update_fields=['status', 'resolved_by', 'updated_at'])
        return Response({'status': 'resolved'})

    @action(detail=False, methods=['post'], url_path='auto')
    def auto_send(self, request):
        """Bulk-create queries for excessive lateness/absence.

        Body:
        - kind: "LATE" | "ABSENT" (required)
        - year_month: "YYYYMM" (optional, defaults to previous month)
        """

        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        kind = (request.data or {}).get('kind')
        kind = (kind or '').upper().strip()
        if kind not in ('LATE', 'ABSENT'):
            return Response({'detail': 'kind must be LATE or ABSENT'}, status=400)

        ym = (request.data or {}).get('year_month')
        if ym and (not str(ym).isdigit() or len(str(ym)) != 6):
            return Response({'detail': 'year_month must be YYYYMM'}, status=400)

        start, end = _prev_month_bounds()
        if ym:
            y = int(str(ym)[:4])
            m = int(str(ym)[4:])
            start = timezone.datetime(y, m, 1).date()
            # end = last day of month
            nxt = (start.replace(day=28) + timezone.timedelta(days=4)).replace(day=1)
            end = nxt - timezone.timedelta(days=1)
        ym = start.strftime('%Y%m')

        # Determine scope
        org_user = user
        branch = None
        if role == 'BRANCH':
            branch = BranchOffice.objects.filter(admin=user).first()
            if not branch:
                return Response({'created': 0, 'skipped': 0, 'detail': 'No branch assigned'})
            org_user = branch.user

        prof = OrganizationProfile.objects.filter(user=org_user).first()
        late_time = getattr(prof, 'late_time', None)
        max_absent = getattr(prof, 'max_days_absent', None)
        max_late = getattr(prof, 'max_days_late', None)

        qs = CorpMember.objects.filter(user=org_user)
        if branch:
            qs = qs.filter(branch=branch)

        # preload logs
        acc_ids = list(qs.values_list('account_id', flat=True))
        logs = AttendanceLog.objects.filter(account_id__in=acc_ids, date__gte=start, date__lte=end)
        logs_by_acc = {}
        for lg in logs:
            logs_by_acc.setdefault(lg.account_id, []).append(lg)

        created = 0
        skipped = 0
        for cm in qs.select_related('branch'):
            cm_logs = logs_by_acc.get(getattr(cm, 'account_id', None), [])
            work_days = _working_days_for_corper(cm, start, end)
            work_set = set(work_days)
            all_present_dates = set([lg.date for lg in cm_logs])
            present_required = len(all_present_dates & work_set)

            cds_day = getattr(cm, 'cds_day', None)
            present_cds = 0
            try:
                cds_int = int(cds_day) if cds_day is not None else None
            except Exception:
                cds_int = None
            if cds_int is not None and 0 <= cds_int <= 4:
                present_cds = len([d for d in all_present_dates if d.weekday() == cds_int])
            present = min(len(work_days), present_required + present_cds)

            late = 0
            if late_time:
                for lg in cm_logs:
                    if lg.date in work_set and lg.time_in and lg.time_in > late_time:
                        late += 1
            absent = max(0, len(work_days) - present)

            if kind == 'LATE':
                if max_late is None or late <= (max_late or 0):
                    skipped += 1
                    continue
                title = f'Attendance Query: Excessive lateness ({ym})'
                message = f'{cm.full_name} ({cm.state_code}) recorded {late} late day(s) in {ym}. Limit: {max_late}.'
            else:
                if max_absent is None or absent <= (max_absent or 0):
                    skipped += 1
                    continue
                title = f'Attendance Query: Excessive absence ({ym})'
                message = f'{cm.full_name} ({cm.state_code}) recorded {absent} absent day(s) in {ym}. Limit: {max_absent}.'

            # avoid duplicates for the month/kind
            if QueryRecord.objects.filter(org=org_user, corper=cm, title=title).exists():
                skipped += 1
                continue

            QueryRecord.objects.create(
                org=org_user,
                branch=getattr(cm, 'branch', None),
                corper=cm,
                title=title,
                message=message,
                status='OPEN',
                created_by=user,
            )
            created += 1

        return Response({'created': created, 'skipped': skipped, 'year_month': ym})


class AttendanceReportView(APIView):
    """Basic attendance report for ORG/BRANCH.

    Returns JSON by default, or CSV when `format=csv`.
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        # Determine org scope
        org_user = user
        branch = None
        if role == 'BRANCH':
            branch = BranchOffice.objects.filter(admin=user).first()
            if not branch:
                return Response({'rows': [], 'summary': {}}, status=200)
            org_user = branch.user

        try:
            start = timezone.datetime.fromisoformat(request.query_params.get('start')).date() if request.query_params.get('start') else None
        except Exception:
            start = None
        try:
            end = timezone.datetime.fromisoformat(request.query_params.get('end')).date() if request.query_params.get('end') else None
        except Exception:
            end = None
        if not end:
            end = timezone.localdate()
        if not start:
            start = end - timezone.timedelta(days=29)
        if start > end:
            start, end = end, start

        qs = AttendanceLog.objects.filter(org=org_user, date__gte=start, date__lte=end)
        if branch:
            # Scope to corpers under this branch (via corp member mapping)
            acc_ids = list(CorpMember.objects.filter(branch=branch).values_list('account_id', flat=True))
            qs = qs.filter(account_id__in=acc_ids)

        rows = []
        total_checkins = 0
        for i in range((end - start).days + 1):
            day = start + timezone.timedelta(days=i)
            day_qs = qs.filter(date=day)
            count = day_qs.count()
            total_checkins += count
            # hours
            hours = 0.0
            from datetime import datetime
            for log in day_qs:
                if log.time_in and log.time_out:
                    try:
                        start_dt = datetime.combine(day, log.time_in)
                        end_dt = datetime.combine(day, log.time_out)
                        delta = end_dt - start_dt
                        if delta.total_seconds() > 0:
                            hours += delta.total_seconds() / 3600.0
                    except Exception:
                        pass
            rows.append({'date': day.isoformat(), 'checkins': count, 'hours': round(hours, 2)})

        summary = {
            'start': start.isoformat(),
            'end': end.isoformat(),
            'days': len(rows),
            'total_checkins': total_checkins,
            'total_hours': round(sum(r['hours'] for r in rows), 2),
        }

        if request.query_params.get('format') == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="attendance_report_{start.isoformat()}_{end.isoformat()}.csv"'
            w = csv.writer(resp)
            w.writerow(['date', 'checkins', 'hours'])
            for r in rows:
                w.writerow([r['date'], r['checkins'], r['hours']])
            return resp

        return Response({'summary': summary, 'rows': rows})


class CorperAttendanceReportView(APIView):
    """Per-corper attendance report for ORG/BRANCH."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        org_user = user
        branch = None
        if role == 'BRANCH':
            branch = BranchOffice.objects.filter(admin=user).first()
            if not branch:
                return Response({'rows': [], 'summary': {}}, status=200)
            org_user = branch.user

        try:
            start = timezone.datetime.fromisoformat(request.query_params.get('start')).date() if request.query_params.get('start') else None
        except Exception:
            start = None
        try:
            end = timezone.datetime.fromisoformat(request.query_params.get('end')).date() if request.query_params.get('end') else None
        except Exception:
            end = None
        if not end:
            end = timezone.localdate()
        if not start:
            start = end - timezone.timedelta(days=29)
        if start > end:
            start, end = end, start

        prof = OrganizationProfile.objects.filter(user=org_user).first()
        late_time = getattr(prof, 'late_time', None)

        cm_qs = CorpMember.objects.filter(user=org_user).select_related('branch', 'department', 'unit', 'account')
        if branch:
            cm_qs = cm_qs.filter(branch=branch)

        acc_ids = list(cm_qs.values_list('account_id', flat=True))
        logs = AttendanceLog.objects.filter(account_id__in=acc_ids, date__gte=start, date__lte=end)
        logs_by_acc = {}
        for lg in logs:
            logs_by_acc.setdefault(lg.account_id, []).append(lg)

        rows = []
        for cm in cm_qs:
            cm_logs = logs_by_acc.get(getattr(cm, 'account_id', None), [])
            work_days = _working_days_for_corper(cm, start, end)
            work_set = set(work_days)
            all_present_dates = set([lg.date for lg in cm_logs])
            present_required = len(all_present_dates & work_set)

            cds_day = getattr(cm, 'cds_day', None)
            present_cds = 0
            try:
                cds_int = int(cds_day) if cds_day is not None else None
            except Exception:
                cds_int = None
            if cds_int is not None and 0 <= cds_int <= 4:
                present_cds = len([d for d in all_present_dates if d.weekday() == cds_int])
            present = min(len(work_days), present_required + present_cds)
            absent = max(0, len(work_days) - present)

            late = 0
            if late_time:
                for lg in cm_logs:
                    if lg.date in work_set and lg.time_in and lg.time_in > late_time:
                        late += 1

            hours = 0.0
            from datetime import datetime
            for lg in cm_logs:
                if lg.date in work_set and lg.time_in and lg.time_out:
                    try:
                        start_dt = datetime.combine(lg.date, lg.time_in)
                        end_dt = datetime.combine(lg.date, lg.time_out)
                        delta = end_dt - start_dt
                        if delta.total_seconds() > 0:
                            hours += delta.total_seconds() / 3600.0
                    except Exception:
                        pass

            rows.append({
                'corper_id': cm.id,
                'full_name': cm.full_name,
                'state_code': cm.state_code,
                'email': cm.email,
                'branch': getattr(cm.branch, 'name', ''),
                'department': getattr(cm.department, 'name', ''),
                'unit': getattr(cm.unit, 'name', ''),
                'working_days': len(work_days),
                'present_days': present,
                'absent_days': absent,
                'late_days': late,
                'hours': round(hours, 2),
            })

        summary = {
            'start': start.isoformat(),
            'end': end.isoformat(),
            'count': len(rows),
        }

        if request.query_params.get('format') == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="corper_report_{start.isoformat()}_{end.isoformat()}.csv"'
            w = csv.writer(resp)
            w.writerow(['full_name', 'state_code', 'email', 'branch', 'department', 'unit', 'working_days', 'present_days', 'absent_days', 'late_days', 'hours'])
            for r in rows:
                w.writerow([r['full_name'], r['state_code'], r['email'], r['branch'], r['department'], r['unit'], r['working_days'], r['present_days'], r['absent_days'], r['late_days'], r['hours']])
            return resp

        return Response({'summary': summary, 'rows': rows})


class AttendanceLogReportView(APIView):
    """Row-level attendance logs report (ORG/BRANCH)."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        org_user = user
        branch = None
        if role == 'BRANCH':
            branch = BranchOffice.objects.filter(admin=user).first()
            if not branch:
                return Response({'rows': []}, status=200)
            org_user = branch.user

        try:
            start = timezone.datetime.fromisoformat(request.query_params.get('start')).date() if request.query_params.get('start') else None
        except Exception:
            start = None
        try:
            end = timezone.datetime.fromisoformat(request.query_params.get('end')).date() if request.query_params.get('end') else None
        except Exception:
            end = None
        if not end:
            end = timezone.localdate()
        if not start:
            start = end - timezone.timedelta(days=29)
        if start > end:
            start, end = end, start

        cm_qs = CorpMember.objects.filter(user=org_user).select_related('branch', 'account')
        if branch:
            cm_qs = cm_qs.filter(branch=branch)
        acc_ids = list(cm_qs.values_list('account_id', flat=True))
        qs = AttendanceLog.objects.filter(account_id__in=acc_ids, date__gte=start, date__lte=end).order_by('-date', '-created_at')

        # Map account to corper details
        cm_by_acc = {cm.account_id: cm for cm in cm_qs}
        rows = []
        for lg in qs:
            cm = cm_by_acc.get(lg.account_id)
            rows.append({
                'date': lg.date.isoformat(),
                'time_in': lg.time_in.isoformat() if lg.time_in else None,
                'time_out': lg.time_out.isoformat() if lg.time_out else None,
                'full_name': getattr(cm, 'full_name', lg.name),
                'state_code': getattr(cm, 'state_code', lg.code),
                'branch': getattr(getattr(cm, 'branch', None), 'name', ''),
            })

        if request.query_params.get('format') == 'csv':
            import csv
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="attendance_logs_{start.isoformat()}_{end.isoformat()}.csv"'
            w = csv.writer(resp)
            w.writerow(['date', 'time_in', 'time_out', 'full_name', 'state_code', 'branch'])
            for r in rows:
                w.writerow([r['date'], r['time_in'], r['time_out'], r['full_name'], r['state_code'], r['branch']])
            return resp

        return Response({'rows': rows, 'summary': {'start': start.isoformat(), 'end': end.isoformat(), 'count': len(rows)}})


def _excel_logo_path(profile):
    try:
        if profile and getattr(profile, 'logo', None) and profile.logo.name and os.path.exists(profile.logo.path):
            return profile.logo.path
    except Exception:
        pass
    return None


def _excel_autofit(ws, max_width=42):
    try:
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(ws.columns, 1):
            width = 10
            for cell in col:
                value = cell.value
                if value is not None:
                    width = max(width, min(len(str(value)) + 2, max_width))
            ws.column_dimensions[get_column_letter(idx)].width = width
    except Exception:
        pass


def _excel_style_header(ws, title, profile=None, merge_to='K1'):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(f'A1:{merge_to}')
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='556B2F')
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 28
    logo_path = _excel_logo_path(profile)
    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.height = 52
            img.width = 90
            ws.add_image(img, 'J2')
        except Exception:
            pass


def _excel_label_value(ws, row, label, value):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='EEF2E8')
    ws.cell(row=row, column=2, value=value or '—')
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)


def _excel_table_header(ws, row, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    border = Border(bottom=Side(style='thin', color='D9E1D2'))
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='556B2F')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border


class AttendanceExcelExportView(APIView):
    """Download an Excel workbook with Daily/Corpers/Logs sheets (ORG/BRANCH)."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Not allowed')

        org_user = user
        branch = None
        if role == 'BRANCH':
            branch = BranchOffice.objects.filter(admin=user).first()
            if not branch:
                return Response({'detail': 'No branch assigned'}, status=400)
            org_user = branch.user

        try:
            start = timezone.datetime.fromisoformat(request.query_params.get('start')).date() if request.query_params.get('start') else None
        except Exception:
            start = None
        try:
            end = timezone.datetime.fromisoformat(request.query_params.get('end')).date() if request.query_params.get('end') else None
        except Exception:
            end = None
        if not end:
            end = timezone.localdate()
        if not start:
            start = end - timezone.timedelta(days=29)
        if start > end:
            start, end = end, start

        # Daily summary rows
        daily_qs = AttendanceLog.objects.filter(org=org_user, date__gte=start, date__lte=end)
        if branch:
            acc_ids = list(CorpMember.objects.filter(branch=branch).values_list('account_id', flat=True))
            daily_qs = daily_qs.filter(account_id__in=acc_ids)

        daily_rows = []
        for i in range((end - start).days + 1):
            day = start + timezone.timedelta(days=i)
            day_qs = daily_qs.filter(date=day)
            count = day_qs.count()
            hours = 0.0
            from datetime import datetime
            for log in day_qs:
                if log.time_in and log.time_out:
                    try:
                        start_dt = datetime.combine(day, log.time_in)
                        end_dt = datetime.combine(day, log.time_out)
                        delta = end_dt - start_dt
                        if delta.total_seconds() > 0:
                            hours += delta.total_seconds() / 3600.0
                    except Exception:
                        pass
            daily_rows.append({'date': day.isoformat(), 'checkins': count, 'hours': round(hours, 2)})

        # Per-corper rows
        prof = OrganizationProfile.objects.filter(user=org_user).first()
        late_time = getattr(prof, 'late_time', None)
        cm_qs = CorpMember.objects.filter(user=org_user).select_related('branch', 'department', 'unit', 'account')
        if branch:
            cm_qs = cm_qs.filter(branch=branch)
        acc_ids = list(cm_qs.values_list('account_id', flat=True))
        logs = AttendanceLog.objects.filter(account_id__in=acc_ids, date__gte=start, date__lte=end)
        logs_by_acc = {}
        for lg in logs:
            logs_by_acc.setdefault(lg.account_id, []).append(lg)

        corper_rows = []
        for cm in cm_qs:
            cm_logs = logs_by_acc.get(getattr(cm, 'account_id', None), [])
            work_days = _working_days_for_corper(cm, start, end)
            work_set = set(work_days)
            all_present_dates = set([lg.date for lg in cm_logs])
            present_required = len(all_present_dates & work_set)

            cds_day = getattr(cm, 'cds_day', None)
            present_cds = 0
            try:
                cds_int = int(cds_day) if cds_day is not None else None
            except Exception:
                cds_int = None
            if cds_int is not None and 0 <= cds_int <= 4:
                present_cds = len([d for d in all_present_dates if d.weekday() == cds_int])
            present = min(len(work_days), present_required + present_cds)
            absent = max(0, len(work_days) - present)

            late = 0
            if late_time:
                for lg in cm_logs:
                    if lg.date in work_set and lg.time_in and lg.time_in > late_time:
                        late += 1

            hours = 0.0
            from datetime import datetime
            for lg in cm_logs:
                if lg.date in work_set and lg.time_in and lg.time_out:
                    try:
                        start_dt = datetime.combine(lg.date, lg.time_in)
                        end_dt = datetime.combine(lg.date, lg.time_out)
                        delta = end_dt - start_dt
                        if delta.total_seconds() > 0:
                            hours += delta.total_seconds() / 3600.0
                    except Exception:
                        pass

            corper_rows.append({
                'full_name': cm.full_name,
                'state_code': cm.state_code,
                'email': cm.email,
                'branch': getattr(cm.branch, 'name', ''),
                'department': getattr(cm.department, 'name', ''),
                'unit': getattr(cm.unit, 'name', ''),
                'working_days': len(work_days),
                'present_days': present,
                'absent_days': absent,
                'late_days': late,
                'hours': round(hours, 2),
            })

        # Row-level logs
        cm_qs2 = CorpMember.objects.filter(user=org_user).select_related('branch', 'account')
        if branch:
            cm_qs2 = cm_qs2.filter(branch=branch)
        acc_ids2 = list(cm_qs2.values_list('account_id', flat=True))
        logs_qs = AttendanceLog.objects.filter(account_id__in=acc_ids2, date__gte=start, date__lte=end).order_by('-date', '-created_at')
        cm_by_acc = {cm.account_id: cm for cm in cm_qs2}
        log_rows = []
        for lg in logs_qs:
            cm = cm_by_acc.get(lg.account_id)
            log_rows.append({
                'date': lg.date.isoformat(),
                'time_in': lg.time_in.isoformat() if lg.time_in else None,
                'time_out': lg.time_out.isoformat() if lg.time_out else None,
                'full_name': getattr(cm, 'full_name', lg.name),
                'state_code': getattr(cm, 'state_code', lg.code),
                'branch': getattr(getattr(cm, 'branch', None), 'name', ''),
            })

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception:
            return Response({'detail': 'Excel export unavailable (missing openpyxl)'}, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'

        scope_name = getattr(branch, 'name', '') if branch else 'All branches'
        generated_at = timezone.localtime().strftime('%Y-%m-%d %H:%M')
        total_checkins = sum(int(r.get('checkins') or 0) for r in daily_rows)
        total_hours = sum(Decimal(str(r.get('hours') or 0)) for r in daily_rows)
        total_absent = sum(int(r.get('absent_days') or 0) for r in corper_rows)
        total_late = sum(int(r.get('late_days') or 0) for r in corper_rows)

        _excel_style_header(ws, 'Attendance Report', prof)
        details = [
            ('Organization', getattr(org_user, 'name', '') or getattr(org_user, 'email', '')),
            ('Email', getattr(org_user, 'email', '')),
            ('Phone', getattr(org_user, 'phone_number', '')),
            ('Address', getattr(org_user, 'address', '')),
            ('Scope', scope_name),
            ('Period', f'{start.isoformat()} to {end.isoformat()}'),
            ('Generated', generated_at),
        ]
        for offset, (label, value) in enumerate(details, 3):
            _excel_label_value(ws, offset, label, value)
        ws['A12'] = 'Summary'
        ws['A12'].font = Font(bold=True, size=13)
        summary_items = [
            ('Days in range', len(daily_rows)),
            ('Total check-ins', total_checkins),
            ('Total hours', float(total_hours)),
            ('Corpers listed', len(corper_rows)),
            ('Attendance log rows', len(log_rows)),
            ('Total absent days', total_absent),
            ('Total late days', total_late),
        ]
        for idx, (label, value) in enumerate(summary_items, 13):
            _excel_label_value(ws, idx, label, value)
        _excel_autofit(ws)

        ws_daily = wb.create_sheet('Daily Check-ins')
        _excel_table_header(ws_daily, 1, ['date', 'checkins'])
        for r in daily_rows:
            ws_daily.append([r['date'], r['checkins']])
        ws_daily.freeze_panes = 'A2'
        _excel_autofit(ws_daily)

        ws2 = wb.create_sheet('Corpers')
        _excel_table_header(ws2, 1, ['full_name', 'state_code', 'email', 'branch', 'department', 'unit', 'working_days', 'present_days', 'absent_days', 'late_days', 'hours'])
        for r in corper_rows:
            ws2.append([r['full_name'], r['state_code'], r['email'], r['branch'], r['department'], r['unit'], r['working_days'], r['present_days'], r['absent_days'], r['late_days'], r['hours']])
        ws2.freeze_panes = 'A2'
        _excel_autofit(ws2)

        ws3 = wb.create_sheet('Logs')
        _excel_table_header(ws3, 1, ['date', 'time_in', 'time_out', 'full_name', 'state_code', 'branch'])
        for r in log_rows:
            ws3.append([r['date'], r['time_in'], r['time_out'], r['full_name'], r['state_code'], r['branch']])
        ws3.freeze_panes = 'A2'
        _excel_autofit(ws3)

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="attendance_report_{start.isoformat()}_{end.isoformat()}.xlsx"'
        return resp


from decimal import Decimal

VAT_RATE = Decimal('0.075')  # 7.5%
CLEARANCE_FEE = Decimal('300.00')


def _format_ngn(amount):
    try:
        value = Decimal(str(amount or '0')).quantize(Decimal('0.01'))
    except Exception:
        value = Decimal('0.00')
    return f"₦{value:,.2f}"


def _ensure_wallet_with_welcome(user):
    acct, created = WalletAccount.objects.get_or_create(user=user)
    # Apply welcome bonus only for organization users
    if getattr(user, 'role', None) == 'ORG' and (created or not acct.transactions.exists()):
        from .models import SystemSetting
        settings = SystemSetting.current()
        welcome = settings.welcome_bonus or Decimal('0.00')
        if welcome > 0:
            vat = Decimal('0.00')
            total = welcome
            WalletTransaction.objects.create(
                account=acct,
                type='CREDIT',
                amount=welcome,
                vat_amount=vat,
                total_amount=total,
                description='Welcome bonus',
                reference='WELCOME'
            )
            acct.balance = (acct.balance or Decimal('0.00')) + total
            acct.save(update_fields=['balance'])
    return acct


def _clearance_year_month(reference):
    ref = str(reference or '')
    suffix = ref.rsplit('-', 1)[-1]
    return suffix if len(suffix) == 6 and suffix.isdigit() else timezone.localdate().strftime('%Y%m')


def _record_clearance_access(cm, reference, source):
    return ClearanceAccess.objects.get_or_create(
        corper=cm,
        year_month=_clearance_year_month(reference),
        defaults={
            'org': cm.user,
            'branch': cm.branch,
            'reference': str(reference or '')[:64],
            'source': source,
        }
    )


def _clearance_access_exists(cm, reference=None):
    qs = ClearanceAccess.objects.filter(corper=cm)
    if reference:
        return qs.filter(reference=str(reference)[:64]).exists()
    return qs.exists()


def _org_subscription_clearance_status(org_user):
    subscription = OrganizationSubscription.objects.filter(org=org_user).select_related('plan').first()
    if not subscription:
        return False, 'the organization does not have an active subscription'
    if subscription.status != 'ACTIVE':
        return False, f"the organization subscription is {subscription.status.lower()}"
    if subscription.expires_at and subscription.expires_at < timezone.now():
        subscription.status = 'EXPIRED'
        subscription.save(update_fields=['status', 'updated_at'])
        return False, 'the organization subscription has expired'
    plan = subscription.plan
    if not plan or not plan.is_active:
        return False, 'the organization subscription plan is not active'
    corper_count = CorpMember.objects.filter(user=org_user).count()
    if plan.corper_max is not None and corper_count > plan.corper_max:
        return False, (
            f"the organization has {corper_count} corpers, which exceeds the "
            f"{plan.name} plan limit of {plan.corper_max}"
        )
    return True, f"covered by the active {subscription.plan_name} subscription"


def _clearance_charge_amount():
    from .models import SystemSetting
    settings_obj = SystemSetting.current()
    try:
        amount = settings_obj.clearance_fee or CLEARANCE_FEE
    except Exception:
        amount = CLEARANCE_FEE
    try:
        if getattr(settings_obj, 'discount_enabled', False):
            pct = Decimal(str(settings_obj.discount_percent or '0'))
            if pct > 0:
                amount = (amount * (Decimal('100') - pct) / Decimal('100')).quantize(Decimal('0.01'))
    except Exception:
        pass
    vat = (amount * VAT_RATE).quantize(Decimal('0.01'))
    total = amount + vat
    return amount, vat, total


def _authorize_clearance_access(cm, corper_user, reference, debit_label='Clearance download charge'):
    ref = str(reference or '')[:64]
    if not ref:
        return False, 'Clearance reference is missing. Please refresh and try again.', None
    if _clearance_access_exists(cm, ref):
        return True, 'Clearance access already recorded.', 'EXISTING'
    if WalletTransaction.objects.filter(reference=ref, type='DEBIT').exists():
        _record_clearance_access(cm, ref, 'EXISTING_WALLET_CHARGE')
        return True, 'Clearance access already paid.', 'EXISTING_WALLET_CHARGE'

    subscription_ok, subscription_reason = _org_subscription_clearance_status(cm.user)
    if subscription_ok:
        _record_clearance_access(cm, ref, 'SUBSCRIPTION')
        return True, subscription_reason, 'SUBSCRIPTION'

    amount, vat, total = _clearance_charge_amount()
    branch_user = getattr(cm.branch, 'admin', None)
    debit_order = (
        (cm.user, 'ORG_WALLET', f'{debit_label} (org)'),
        (branch_user, 'BRANCH_WALLET', f'{debit_label} (admin)'),
        (corper_user, 'CORPER_WALLET', f'{debit_label} (corper)'),
    )

    for user, source, description in debit_order:
        if not user:
            continue
        acct = _ensure_wallet_with_welcome(user)
        if (acct.balance or Decimal('0.00')) >= total:
            WalletTransaction.objects.create(
                account=acct,
                type='DEBIT',
                amount=amount,
                vat_amount=vat,
                total_amount=total,
                description=description,
                reference=ref
            )
            acct.balance = acct.balance - total
            acct.save(update_fields=['balance'])
            _record_clearance_access(cm, ref, source)
            return True, f'Charged {description.lower()} successfully.', source

    reason = (
        f"Clearance is not covered by subscription because {subscription_reason}. "
        f"Wallet fallback also failed because no organization, admin, or corper wallet "
        f"has enough available balance for the clearance fee of {_format_ngn(total)}."
    )
    return False, reason, None


class WalletView(APIView):
    def get(self, request):
        user = request.user
        # Allow ORG, BRANCH, CORPER to view their own wallet
        if getattr(user, 'role', None) not in ('ORG', 'BRANCH', 'CORPER'):
            raise PermissionDenied('Not allowed')
        acct = _ensure_wallet_with_welcome(user)
        from .serializers import WalletAccountSerializer
        data = WalletAccountSerializer(acct).data
        return Response(data)


class WalletStatementExportView(APIView):
    """Download the authenticated user's complete wallet statement as Excel."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        if getattr(user, 'role', None) not in ('ORG', 'BRANCH', 'CORPER'):
            raise PermissionDenied('Not allowed')

        acct = _ensure_wallet_with_welcome(user)
        txs = list(acct.transactions.all().order_by('created_at', 'id'))
        total_credit = sum((t.total_amount or Decimal('0.00')) for t in txs if t.type == 'CREDIT')
        total_debit = sum((t.total_amount or Decimal('0.00')) for t in txs if t.type == 'DEBIT')

        org_user = user
        branch_name = ''
        try:
            if getattr(user, 'role', None) == 'BRANCH':
                branch = BranchOffice.objects.filter(admin=user).select_related('user').first()
                if branch:
                    org_user = branch.user
                    branch_name = branch.name
            elif getattr(user, 'role', None) == 'CORPER':
                cm = getattr(user, 'corper_profile', None)
                if cm:
                    org_user = cm.user
                    branch_name = getattr(cm.branch, 'name', '') if cm.branch else ''
        except Exception:
            pass
        profile = OrganizationProfile.objects.filter(user=org_user).first()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            return Response({'detail': 'Excel export unavailable (missing openpyxl)'}, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Wallet Statement'
        _excel_style_header(ws, 'Wallet Statement', profile, merge_to='I1')

        generated_at = timezone.localtime().strftime('%Y-%m-%d %H:%M')
        account_details = [
            ('Organization', getattr(org_user, 'name', '') or getattr(org_user, 'email', '')),
            ('Organization Email', getattr(org_user, 'email', '')),
            ('Organization Phone', getattr(org_user, 'phone_number', '')),
            ('Organization Address', getattr(org_user, 'address', '')),
            ('Statement Account', getattr(user, 'name', '') or getattr(user, 'email', '')),
            ('Account Email', getattr(user, 'email', '')),
            ('Role', getattr(user, 'role', '')),
            ('Branch', branch_name),
            ('Generated', generated_at),
        ]
        for offset, (label, value) in enumerate(account_details, 3):
            _excel_label_value(ws, offset, label, value)

        ws['A14'] = 'Summary'
        ws['A14'].font = Font(bold=True, size=13)
        summary_rows = [
            ('Current balance', float(acct.balance or Decimal('0.00'))),
            ('Total credit', float(total_credit)),
            ('Total debit', float(total_debit)),
        ]
        ws['A15'] = 'Metric'
        ws['B15'] = 'Currency'
        ws['C15'] = 'Amount'
        for cell in ws[15]:
            cell.font = Font(bold=True)
        for offset, (label, value) in enumerate(summary_rows, 16):
            ws.cell(row=offset, column=1, value=label)
            ws.cell(row=offset, column=2, value='NGN')
            amount_cell = ws.cell(row=offset, column=3, value=value)
            amount_cell.number_format = '#,##0.00'
        ws.cell(row=19, column=1, value='Transactions')
        ws.cell(row=19, column=3, value=len(txs)).number_format = '#,##0'

        start_row = 21
        _excel_table_header(ws, start_row, ['date', 'description', 'type', 'currency', 'amount', 'vat', 'total', 'reference'])
        for t in txs:
            ws.append([
                timezone.localtime(t.created_at).strftime('%Y-%m-%d %H:%M') if t.created_at else '',
                t.description,
                t.type,
                'NGN',
                float(t.amount or Decimal('0.00')),
                float(t.vat_amount or Decimal('0.00')),
                float(t.total_amount or Decimal('0.00')),
                t.reference,
            ])
        if txs:
            for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + len(txs), min_col=5, max_col=7):
                for cell in row:
                    cell.number_format = '#,##0.00'
            table = Table(displayName='WalletTransactions', ref=f'A{start_row}:H{start_row + len(txs)}')
            table.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium4',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        protection_password = 'nysc-statement-readonly'
        ws.protection.set_password(protection_password)
        ws.protection.enable()
        wb.security.lockStructure = True
        wb.security.set_workbook_password(protection_password)
        _excel_autofit(ws)

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="wallet_statement_{timezone.localdate().isoformat()}.xlsx"'
        return resp


def wallet_charge_clearance(request):
    """Charge organization when a corper downloads clearance letter.

    Expects POST with JSON: { "reference": "NYSC-..." }
    """
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    user = request.user
    if getattr(user, 'role', None) != 'CORPER':
        return JsonResponse({'detail': 'Only corper can trigger charge'}, status=403)
    try:
        cm = user.corper_profile
    except Exception:
        return JsonResponse({'detail': 'Corper profile not found'}, status=404)
    import json
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        payload = {}
    ref = payload.get('reference', '')
    ok, reason, source = _authorize_clearance_access(
        cm,
        request.user,
        ref,
        debit_label='Clearance download charge',
    )
    if ok:
        status_value = 'covered' if source == 'SUBSCRIPTION' else 'charged'
        return JsonResponse({'status': status_value, 'source': source, 'detail': reason})
    return JsonResponse({'status': 'insufficient', 'detail': reason}, status=402)


class AnnouncementView(APIView):
    def get(self, request):
        user = request.user
        # Only organization dashboard shows this floating announcement
        if getattr(user, 'role', None) != 'ORG':
            return Response(status=204)
        from .models import SystemSetting
        from django.utils import timezone
        s = SystemSetting.current()
        if not getattr(s, 'notify_enabled', False):
            return Response(status=204)
        now = timezone.now()
        if s.notify_start and now < s.notify_start:
            return Response(status=204)
        if s.notify_end and now > s.notify_end:
            return Response(status=204)
        data = {
            'title': s.notify_title or 'Notice',
            'message': s.notify_message or '',
        }
        return Response(data)


class ClearanceStatusView(APIView):
    """List corpers with clearance qualification and download status for previous month.

    - ORG: all corpers under organization
    - BRANCH: corpers under managed branch
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Only organization or branch admins can view clearance status')
        if role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if not b:
                return Response([], status=200)
            corpers = CorpMember.objects.filter(branch=b).select_related('user', 'branch')
            org_user = b.user
        else:
            corpers = CorpMember.objects.filter(user=user).select_related('user', 'branch')
            org_user = user
        start, end = _prev_month_bounds()
        yyyymm = start.strftime('%Y%m')
        prof = OrganizationProfile.objects.filter(user=org_user).first()
        late_time = getattr(prof, 'late_time', None)
        max_absent = getattr(prof, 'max_days_absent', None)
        max_late = getattr(prof, 'max_days_late', None)

        # Preload attendance logs for the window
        acc_ids = list(corpers.values_list('account_id', flat=True))
        logs = AttendanceLog.objects.filter(account_id__in=acc_ids, date__gte=start, date__lte=end)
        logs_by_acc = {}
        for lg in logs:
            logs_by_acc.setdefault(lg.account_id, []).append(lg)

        results = []
        for cm in corpers:
            work_days = _working_days_for_corper(cm, start, end)
            work_set = set(work_days)
            cm_logs = logs_by_acc.get(getattr(cm, 'account_id', None), [])
            all_present_dates = set([lg.date for lg in cm_logs])
            present_required = len(all_present_dates & work_set)
            # CDS days are excluded from required days, but attendance on CDS can count toward the threshold.
            cds_day = getattr(cm, 'cds_day', None)
            present_cds = 0
            try:
                cds_int = int(cds_day) if cds_day is not None else None
            except Exception:
                cds_int = None
            if cds_int is not None and 0 <= cds_int <= 4:
                present_cds = len([d for d in all_present_dates if d.weekday() == cds_int])
            present = min(len(work_days), present_required + present_cds)
            late = 0
            if late_time:
                for lg in cm_logs:
                    # Late applies only on working days (CDS excluded)
                    if lg.date in work_set and lg.time_in and lg.time_in > late_time:
                        late += 1
            absent = max(0, len(work_days) - present)
            is_first = (
                not _clearance_access_exists(cm) and
                not WalletTransaction.objects.filter(type='DEBIT', reference__startswith=f"NYSC-{cm.state_code}-").exists()
            )
            # If corper was enrolled after the clearance month started, do not penalize them for missing days.
            enrolled_after_start = False
            try:
                joined = getattr(cm.account, 'date_joined', None)
                enrolled_after_start = bool(joined and joined.date() > start)
            except Exception:
                enrolled_after_start = False
            override = ClearanceOverride.objects.filter(corper=cm, year_month=yyyymm).exists()
            exceeded_abs = (max_absent is not None and absent > (max_absent or 0))
            exceeded_late = (max_late is not None and late > (max_late or 0))
            qualified = is_first or override or enrolled_after_start or (not exceeded_abs and not exceeded_late)
            reference = f"NYSC-{cm.state_code}-{yyyymm}"
            downloaded = (
                ClearanceAccess.objects.filter(corper=cm, reference=reference[:64]).exists() or
                WalletTransaction.objects.filter(type='DEBIT', reference=reference).exists()
            )
            results.append({
                'id': cm.id,
                'full_name': cm.full_name,
                'state_code': cm.state_code,
                'branch': getattr(cm.branch, 'name', ''),
                'absent': absent,
                'late': late,
                'qualified': qualified,
                'downloaded': downloaded,
                'override': override,
                'reference': reference,
            })
        return Response(results)


class ClearanceApproveView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        role = getattr(user, 'role', None)
        if role not in ('ORG', 'BRANCH'):
            raise PermissionDenied('Only organization or branch admins can approve clearance')
        cm_id = (request.data or {}).get('corper')
        if not cm_id:
            return Response({'detail': 'corper id required'}, status=400)
        cm = CorpMember.objects.filter(id=cm_id).select_related('branch').first()
        if not cm:
            return Response({'detail': 'corper not found'}, status=404)
        if role == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            if not b or cm.branch_id != b.id:
                raise PermissionDenied('Corper does not belong to your branch')
        elif role == 'ORG':
            if cm.user_id != user.id:
                raise PermissionDenied('Corper does not belong to your organization')
        start, end = _prev_month_bounds()
        yyyymm = start.strftime('%Y%m')
        obj, _ = ClearanceOverride.objects.get_or_create(corper=cm, year_month=yyyymm, defaults={'created_by': user})
        return Response({'status': 'approved', 'corper': cm.id, 'year_month': yyyymm})


class WalletFundView(APIView):
    def post(self, request):
        user = request.user
        # Allow only org, or branch admin acting for their org
        if getattr(user, 'role', None) == 'BRANCH':
            b = BranchOffice.objects.filter(admin=user).first()
            user = b.user if b else user
        if getattr(user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organization can fund wallet')
        # Placeholder legacy endpoint
        return Response({'detail': 'Use /api/auth/wallet/paystack/initialize and verify'}, status=202)


SUBSCRIPTION_PLAN_DEFAULTS = [
    {
        'code': 'STARTER',
        'name': 'Starter',
        'corper_min': 0,
        'corper_max': 10,
        'monthly_price': Decimal('5000.00'),
        'yearly_price': Decimal('55000.00'),
        'sort_order': 1,
    },
    {
        'code': 'BASIC',
        'name': 'Basic',
        'corper_min': 10,
        'corper_max': 50,
        'monthly_price': Decimal('25000.00'),
        'yearly_price': Decimal('258000.00'),
        'sort_order': 2,
    },
    {
        'code': 'PRO',
        'name': 'Premium',
        'corper_min': 50,
        'corper_max': 100,
        'monthly_price': Decimal('45000.00'),
        'yearly_price': Decimal('510000.00'),
        'sort_order': 3,
    },
    {
        'code': 'ENTERPRISE',
        'name': 'Enterprise',
        'corper_min': 100,
        'corper_max': None,
        'monthly_price': Decimal('95000.00'),
        'yearly_price': Decimal('999999.00'),
        'sort_order': 4,
    },
]


def _ensure_subscription_plans():
    for cfg in SUBSCRIPTION_PLAN_DEFAULTS:
        SubscriptionPlanSetting.objects.get_or_create(code=cfg['code'], defaults=cfg)


def _money_value(value):
    try:
        return Decimal(str(value or '0')).quantize(Decimal('0.01'))
    except Exception:
        return Decimal('0.00')


def _discounted_subscription_amount(plan, billing_cycle):
    original = _money_value(plan.yearly_price if billing_cycle == 'YEARLY' else plan.monthly_price)
    discount_amount = Decimal('0.00')
    if getattr(plan, 'discount_enabled', False):
        try:
            pct = Decimal(str(plan.discount_percent or '0'))
            if pct > 0:
                discount_amount = (original * pct / Decimal('100')).quantize(Decimal('0.01'))
        except Exception:
            discount_amount = Decimal('0.00')
    charged = max(Decimal('0.00'), (original - discount_amount).quantize(Decimal('0.01')))
    return original, discount_amount, charged


def _plan_range_label(plan):
    if plan.corper_max is None:
        return f"{plan.corper_min}+ corpers"
    return f"{plan.corper_min}-{plan.corper_max} corpers"


def _subscription_plan_payload(plan):
    monthly_original, monthly_discount, monthly_charged = _discounted_subscription_amount(plan, 'MONTHLY')
    yearly_original, yearly_discount, yearly_charged = _discounted_subscription_amount(plan, 'YEARLY')
    custom_pricing = str(plan.code or '').upper() == 'ENTERPRISE' and monthly_original <= 0 and yearly_original <= 0
    return {
        'id': plan.id,
        'code': plan.code,
        'name': plan.name,
        'corper_min': plan.corper_min,
        'corper_max': plan.corper_max,
        'range_label': _plan_range_label(plan),
        'monthly_price': str(monthly_charged),
        'yearly_price': str(yearly_charged),
        'original_monthly_price': str(monthly_original),
        'original_yearly_price': str(yearly_original),
        'monthly_discount_amount': str(monthly_discount),
        'yearly_discount_amount': str(yearly_discount),
        'discount_enabled': bool(plan.discount_enabled),
        'discount_percent': str(plan.discount_percent or Decimal('0.00')),
        'custom_pricing': custom_pricing,
        'is_active': bool(plan.is_active),
    }


def _subscription_payload(subscription):
    if not subscription:
        return None
    return {
        'plan_code': subscription.plan_code,
        'plan_name': subscription.plan_name,
        'billing_cycle': subscription.billing_cycle,
        'status': subscription.status,
        'amount_paid': str(subscription.amount_paid),
        'starts_at': subscription.starts_at.isoformat() if subscription.starts_at else None,
        'expires_at': subscription.expires_at.isoformat() if subscription.expires_at else None,
    }


def _subscription_payment_payload(payment):
    return {
        'id': payment.id,
        'plan_code': payment.plan_code,
        'plan_name': payment.plan_name,
        'billing_cycle': payment.billing_cycle,
        'amount': str(payment.amount),
        'discount_amount': str(payment.discount_amount),
        'amount_charged': str(payment.amount_charged),
        'reference': payment.reference,
        'status': payment.status,
        'paid_at': payment.paid_at.isoformat() if payment.paid_at else None,
        'created_at': payment.created_at.isoformat() if payment.created_at else None,
    }


def _activate_subscription(payment, paid_at=None):
    now = paid_at or timezone.now()
    duration = timedelta(days=365 if payment.billing_cycle == 'YEARLY' else 30)
    subscription, _ = OrganizationSubscription.objects.update_or_create(
        org=payment.org,
        defaults={
            'plan': payment.plan,
            'plan_code': payment.plan_code,
            'plan_name': payment.plan_name,
            'billing_cycle': payment.billing_cycle,
            'status': 'ACTIVE',
            'amount_paid': payment.amount_charged,
            'starts_at': now,
            'expires_at': now + duration,
        }
    )
    return subscription


class SubscriptionPlansView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        _ensure_subscription_plans()
        plans = SubscriptionPlanSetting.objects.filter(is_active=True).order_by('sort_order', 'id')
        return Response({'plans': [_subscription_plan_payload(plan) for plan in plans]})


class SubscriptionStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organizations can manage subscription')
        _ensure_subscription_plans()
        current = OrganizationSubscription.objects.filter(org=request.user).select_related('plan').first()
        if current and current.status == 'ACTIVE' and current.expires_at and current.expires_at < timezone.now():
            current.status = 'EXPIRED'
            current.save(update_fields=['status', 'updated_at'])
        payments = SubscriptionPayment.objects.filter(org=request.user).select_related('plan')[:50]
        plans = SubscriptionPlanSetting.objects.filter(is_active=True).order_by('sort_order', 'id')
        return Response({
            'current': _subscription_payload(current),
            'payments': [_subscription_payment_payload(payment) for payment in payments],
            'plans': [_subscription_plan_payload(plan) for plan in plans],
        })


class SubscriptionInitializeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organizations can start a subscription')
        _ensure_subscription_plans()
        data = request.data or {}
        plan_code = str(data.get('plan') or '').strip().upper()
        billing_cycle = str(data.get('billing_cycle') or data.get('cycle') or 'MONTHLY').strip().upper()
        if billing_cycle not in ('MONTHLY', 'YEARLY'):
            return Response({'detail': 'billing_cycle must be MONTHLY or YEARLY'}, status=400)
        plan = SubscriptionPlanSetting.objects.filter(code=plan_code, is_active=True).first()
        if not plan:
            return Response({'detail': 'Subscription plan not found'}, status=404)

        original, discount_amount, charged = _discounted_subscription_amount(plan, billing_cycle)
        if str(plan.code or '').upper() == 'ENTERPRISE' and original <= 0:
            return Response({'detail': 'Contact us for Enterprise pricing at admin@sahabs.tech or +2347082505053.'}, status=400)
        reference = f"SUB-{request.user.id}-{uuid.uuid4().hex[:18]}".upper()
        payment = SubscriptionPayment.objects.create(
            org=request.user,
            plan=plan,
            plan_code=plan.code,
            plan_name=plan.name,
            billing_cycle=billing_cycle,
            amount=original,
            discount_amount=discount_amount,
            amount_charged=charged,
            reference=reference,
            status='PENDING',
        )
        if charged <= 0:
            payment.status = 'SUCCESS'
            payment.paid_at = timezone.now()
            payment.raw_response = {'free_plan': True}
            payment.save(update_fields=['status', 'paid_at', 'raw_response', 'updated_at'])
            subscription = _activate_subscription(payment, payment.paid_at)
            return Response({
                'status': 'success',
                'free': True,
                'reference': payment.reference,
                'subscription': _subscription_payload(subscription),
            })

        try:
            from .utils import get_paystack_keys
            keys = get_paystack_keys()
        except Exception as e:
            payment.status = 'FAILED'
            payment.raw_response = {'error': str(e)}
            payment.save(update_fields=['status', 'raw_response', 'updated_at'])
            return Response({'detail': str(e)}, status=400)

        payload = {
            'email': getattr(request.user, 'email', ''),
            'amount': int(charged * Decimal('100')),
            'reference': payment.reference,
            'metadata': {
                'uid': int(getattr(request.user, 'id', 0) or 0),
                'role': getattr(request.user, 'role', '') or '',
                'email': getattr(request.user, 'email', '') or '',
                'purpose': 'subscription',
                'subscription_payment_id': payment.id,
                'plan': plan.code,
                'billing_cycle': billing_cycle,
            },
        }
        cb = data.get('callback_url')
        if cb:
            payload['callback_url'] = cb
        headers = {
            'Authorization': f"Bearer {keys['secret_key']}",
            'Content-Type': 'application/json'
        }
        try:
            r = requests.post('https://api.paystack.co/transaction/initialize', json=payload, headers=headers, timeout=20)
            jr = r.json()
        except Exception:
            payment.status = 'FAILED'
            payment.raw_response = {'error': 'Failed to reach Paystack'}
            payment.save(update_fields=['status', 'raw_response', 'updated_at'])
            return Response({'detail': 'Failed to reach Paystack'}, status=502)
        payment.raw_response = jr
        payment.save(update_fields=['raw_response', 'updated_at'])
        if r.status_code != 200 or not jr.get('status'):
            payment.status = 'FAILED'
            payment.save(update_fields=['status', 'updated_at'])
            return Response({'detail': jr.get('message', 'Initialization failed')}, status=400)
        pdata = jr.get('data') or {}
        return Response({
            'authorization_url': pdata.get('authorization_url'),
            'reference': payment.reference,
            'amount': str(payment.amount_charged),
            'plan': _subscription_plan_payload(plan),
        })


class SubscriptionVerifyView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if getattr(request.user, 'role', None) != 'ORG':
            raise PermissionDenied('Only organizations can verify subscription payments')
        ref = (request.data or {}).get('reference')
        if not ref:
            return Response({'detail': 'reference is required'}, status=400)
        payment = SubscriptionPayment.objects.filter(reference=str(ref)[:64], org=request.user).select_related('plan').first()
        if not payment:
            return Response({'detail': 'Subscription payment not found'}, status=404)
        if payment.status == 'SUCCESS':
            subscription = OrganizationSubscription.objects.filter(org=request.user).first()
            return Response({'status': 'success', 'duplicate': True, 'subscription': _subscription_payload(subscription)})

        try:
            from .utils import get_paystack_keys
            keys = get_paystack_keys()
        except Exception as e:
            return Response({'detail': str(e)}, status=400)

        headers = {
            'Authorization': f"Bearer {keys['secret_key']}",
            'Content-Type': 'application/json'
        }
        try:
            r = requests.get(f'https://api.paystack.co/transaction/verify/{payment.reference}', headers=headers, timeout=20)
            jr = r.json()
        except Exception:
            return Response({'detail': 'Failed to reach Paystack'}, status=502)
        if r.status_code != 200 or not jr.get('status'):
            return Response({'detail': jr.get('message', 'Verification failed')}, status=400)
        data = jr.get('data') or {}
        if data.get('status') != 'success':
            payment.status = 'FAILED'
            payment.raw_response = data
            payment.save(update_fields=['status', 'raw_response', 'updated_at'])
            return Response({'status': 'failed'}, status=200)

        paid_kobo = data.get('amount') or 0
        paid = _money_value(Decimal(paid_kobo) / Decimal('100'))
        if paid < payment.amount_charged:
            payment.status = 'FAILED'
            payment.raw_response = data
            payment.save(update_fields=['status', 'raw_response', 'updated_at'])
            return Response({'detail': 'Paid amount is lower than the selected subscription price'}, status=400)

        payment.status = 'SUCCESS'
        payment.raw_response = data
        payment.paid_at = timezone.now()
        payment.save(update_fields=['status', 'raw_response', 'paid_at', 'updated_at'])
        subscription = _activate_subscription(payment, payment.paid_at)
        return Response({'status': 'success', 'subscription': _subscription_payload(subscription)})


class PaystackInitializeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            from .utils import get_paystack_keys
            keys = get_paystack_keys()
        except Exception as e:
            return Response({'detail': str(e)}, status=400)

        data = request.data or {}
        email = data.get('email') or getattr(request.user, 'email', '')
        amount = data.get('amount')
        if not email or not amount:
            return Response({'detail': 'email and amount are required'}, status=400)
        try:
            amt_kobo = int(Decimal(str(amount)) * 100)
        except Exception:
            return Response({'detail': 'invalid amount'}, status=400)
        payload = {
            'email': email,
            'amount': amt_kobo,
        }
        # Include metadata so webhook can reliably credit the correct user later.
        try:
            payload['metadata'] = {
                'uid': int(getattr(request.user, 'id', 0) or 0),
                'role': getattr(request.user, 'role', '') or '',
                'email': getattr(request.user, 'email', '') or '',
                'purpose': 'wallet',
            }
        except Exception:
            pass
        cb = data.get('callback_url')
        if cb:
            payload['callback_url'] = cb
        headers = {
            'Authorization': f"Bearer {keys['secret_key']}",
            'Content-Type': 'application/json'
        }
        try:
            r = requests.post('https://api.paystack.co/transaction/initialize', json=payload, headers=headers, timeout=20)
            jr = r.json()
        except Exception:
            return Response({'detail': 'Failed to reach Paystack'}, status=502)
        if r.status_code != 200 or not jr.get('status'):
            return Response({'detail': jr.get('message', 'Initialization failed')}, status=400)
        pdata = jr.get('data') or {}
        return Response({'authorization_url': pdata.get('authorization_url'), 'reference': pdata.get('reference')})


class PaystackVerifyView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ref = (request.data or {}).get('reference')
        if not ref:
            return Response({'detail': 'reference is required'}, status=400)
        if SubscriptionPayment.objects.filter(reference=str(ref)[:64]).exists():
            return Response({'detail': 'Use the subscription verification endpoint for this reference'}, status=400)
        try:
            from .utils import get_paystack_keys
            keys = get_paystack_keys()
        except Exception as e:
            return Response({'detail': str(e)}, status=400)

        headers = {
            'Authorization': f"Bearer {keys['secret_key']}",
            'Content-Type': 'application/json'
        }
        try:
            r = requests.get(f'https://api.paystack.co/transaction/verify/{ref}', headers=headers, timeout=20)
            jr = r.json()
        except Exception:
            return Response({'detail': 'Failed to reach Paystack'}, status=502)
        if r.status_code != 200 or not jr.get('status'):
            return Response({'detail': jr.get('message', 'Verification failed')}, status=400)
        data = jr.get('data') or {}
        status_val = data.get('status')
        if status_val != 'success':
            return Response({'status': 'failed'}, status=200)
        # Idempotency: don't double-credit for the same reference.
        if WalletTransaction.objects.filter(type='CREDIT', reference=str(ref)[:64]).exists():
            acct = _ensure_wallet_with_welcome(request.user)
            return Response({'status': 'success', 'balance': str(acct.balance), 'duplicate': True})
        # Credit wallet on success using customer email and paid amount
        cust = (data.get('customer') or {})
        email = cust.get('email') or getattr(request.user, 'email', None)
        paid_kobo = data.get('amount') or 0
        try:
            paid = (Decimal(paid_kobo) / Decimal('100')).quantize(Decimal('0.01'))
        except Exception:
            paid = Decimal('0.00')
        if not email or paid <= 0:
            return Response({'detail': 'Missing payment data'}, status=400)
        # Credit the requestor's wallet (ORG/BRANCH/CORPER). Fallback to email match if needed.
        target_user = request.user
        if getattr(target_user, 'role', None) not in ('ORG', 'BRANCH', 'CORPER'):
            from .models import OrganizationUser
            target_user = OrganizationUser.objects.filter(email=email).first() or target_user
        acct = _ensure_wallet_with_welcome(target_user)
        WalletTransaction.objects.create(
            account=acct,
            type='CREDIT',
            amount=paid,
            vat_amount=Decimal('0.00'),
            total_amount=paid,
            description='Wallet funding via Paystack',
            reference=str(ref)[:64]
        )
        acct.balance = (acct.balance or Decimal('0.00')) + paid
        acct.save(update_fields=['balance'])
        return Response({'status': 'success', 'balance': str(acct.balance)})


@method_decorator(csrf_exempt, name='dispatch')
class PaystackWebhookView(APIView):
    """Paystack webhook handler.

    Configure Paystack webhook URL to:
    - https://api.sahabs.tech/api/auth/paystack/webhook/
    """

    authentication_classes = []
    permission_classes = []

    def post(self, request):
        # Validate signature (Paystack uses HMAC-SHA512 of raw body with the secret key).
        sig = request.headers.get('x-paystack-signature') or request.META.get('HTTP_X_PAYSTACK_SIGNATURE')
        if not sig:
            return Response({'detail': 'Missing signature'}, status=400)

        try:
            from .utils import get_paystack_keys
            keys = get_paystack_keys()
        except Exception as e:
            return Response({'detail': str(e)}, status=500)

        key = (keys.get('webhook_secret') or keys.get('secret_key') or '').encode('utf-8')
        if not key:
            return Response({'detail': 'Paystack secret missing'}, status=500)

        body = request.body or b''
        expected = hmac.new(key, body, hashlib.sha512).hexdigest()
        if not hmac.compare_digest(expected, str(sig).strip()):
            return Response({'detail': 'Invalid signature'}, status=401)

        try:
            payload = json.loads(body.decode('utf-8') or '{}')
        except Exception:
            return Response({'detail': 'Invalid JSON'}, status=400)

        event = (payload.get('event') or '').strip()
        data = payload.get('data') or {}
        if event != 'charge.success':
            # Acknowledge other events to avoid retries.
            return Response({'ok': True})

        ref = (data.get('reference') or '').strip()
        if not ref:
            return Response({'detail': 'Missing reference'}, status=400)

        meta = data.get('metadata') or {}
        subscription_payment = SubscriptionPayment.objects.filter(reference=str(ref)[:64]).select_related('plan', 'org').first()
        if subscription_payment or str(meta.get('purpose') or '').lower() == 'subscription':
            if not subscription_payment:
                return Response({'ok': True, 'subscription_unmatched': True})
            if subscription_payment.status == 'SUCCESS':
                return Response({'ok': True, 'subscription_duplicate': True})
            paid_kobo = data.get('amount') or 0
            paid = _money_value(Decimal(paid_kobo) / Decimal('100'))
            if paid < subscription_payment.amount_charged:
                subscription_payment.status = 'FAILED'
                subscription_payment.raw_response = data
                subscription_payment.save(update_fields=['status', 'raw_response', 'updated_at'])
                return Response({'ok': True, 'subscription_invalid_amount': True})
            subscription_payment.status = 'SUCCESS'
            subscription_payment.raw_response = data
            subscription_payment.paid_at = timezone.now()
            subscription_payment.save(update_fields=['status', 'raw_response', 'paid_at', 'updated_at'])
            _activate_subscription(subscription_payment, subscription_payment.paid_at)
            return Response({'ok': True, 'subscription': True})

        # Idempotency: do nothing if already credited.
        if WalletTransaction.objects.filter(type='CREDIT', reference=str(ref)[:64]).exists():
            return Response({'ok': True, 'duplicate': True})

        paid_kobo = data.get('amount') or 0
        try:
            paid = (Decimal(paid_kobo) / Decimal('100')).quantize(Decimal('0.01'))
        except Exception:
            paid = Decimal('0.00')
        if paid <= 0:
            return Response({'detail': 'Invalid amount'}, status=400)

        # Resolve target user: prefer metadata uid, else customer email.
        target_user = None
        uid = meta.get('uid') or meta.get('user_id')
        try:
            uid = int(uid)
        except Exception:
            uid = 0
        if uid:
            try:
                target_user = User.objects.filter(id=uid).first()
            except Exception:
                target_user = None
        if not target_user:
            cust = (data.get('customer') or {})
            email = cust.get('email')
            if email:
                target_user = User.objects.filter(email=email).first()

        if not target_user:
            # Acknowledge to prevent retries; can't map payment to a wallet.
            return Response({'ok': True, 'unmatched': True})

        acct = _ensure_wallet_with_welcome(target_user)
        WalletTransaction.objects.create(
            account=acct,
            type='CREDIT',
            amount=paid,
            vat_amount=Decimal('0.00'),
            total_amount=paid,
            description='Wallet funding via Paystack (webhook)',
            reference=str(ref)[:64]
        )
        acct.balance = (acct.balance or Decimal('0.00')) + paid
        acct.save(update_fields=['balance'])
        return Response({'ok': True})
